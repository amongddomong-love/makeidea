# -*- coding: utf-8 -*-
"""
목적:
- 블룸버그 xbbg로 주요 지표를 조회하고, 1일/10일 변화율(또는 bp/pp) 기반 임계수준 초과 여부를 점검
- 특정 항목은 월평균(MTD/PrevM/3M-ago) 기준으로 스프레드/변동률 임계치 점검
- ⚠️ "현재 값이 나오는 지표만" alerts에 반영 (데이터 미수급 시 해당 블록은 코드에서 주석 처리 예시를 남김)
- 결과를 엑셀 파일(요약 alerts + 원시 raw_data)로 저장

환경 유의:
- Bloomberg Desktop + xbbg (터미널 로그인 상태) 필요
- pip: xbbg, pandas, numpy, openpyxl 등
- Windows 경로는 반드시 pathlib.Path 사용 또는 / 슬래시

임계수준 요약:
- 원화금리(국고3Y): 1일 ±15bp, 10일 ±50bp
- 원화금리(국고10Y): 1일 ±15bp, 10일 ±45bp
- 외화금리(미10Y/Term SOFR 6M): 최근 3개월 평균 대비 ±100bp
- KRWUSD 환율: 1일 ±2%, 10일 ±5%
- KOSPI: 1일 -3.5%, 10일 -10% (하락 방향만)
- VKOSPI: 1일 +5%p, 10일 +10%p (상승 방향만)
- USDKRW 1Y IV: 1일 ±5%p, 10일 ±10%p
- 외화 월평균 장단기 금리차: (SOFR OIS 1Y - TSFR 1M) MTD 평균 ≥ +150bp
- KR 1Y - 기준금리: 5영업일 연속 < -24bp
- 기준금리 - 콜금리: > +40bp (레벨)
- Term SOFR 3M: MTD-PrevM 절대변화 > 75bp
- JPY 3M TIBOR: MTD-PrevM 절대변화 > 25bp
- 한국 5Y CDS: PrevM 대비 +100bp 3영업일 연속 / M-3 대비 +200bp 3영업일 연속
- KR Term Spread (10Y-3Y): 역전(≤0bp) 5영업일 연속
- 국가별 CDS 17개국: 전월 평균 대비 +30% 상승
- (회사채/국고) 3Y 비율: 전월평균 대비 +16% 상승
- 월평균 장단기 (금융채1Y - CD3M): MTD 스프레드 ≥ +70bp
- (금융채1Y AAA - KTB1Y): 5영업일 연속 ≥ +50bp
- S&P500: 1일 ≤ -3%, 10일 ≤ -12%
- EuroStoxx50: 1일 |Δ| ≥ 3%, 10일 ≤ -12%
"""

from pathlib import Path  # ✅ 윈도우에서도 안전한 경로 처리
from datetime import timedelta
import numpy as np
import pandas as pd

# Bloomberg
try:
    from xbbg import blp
except ImportError as e:
    raise SystemExit(
        "xbbg가 설치되어 있지 않습니다. 다음을 먼저 실행해주세요:\n"
        "    pip install xbbg pandas numpy openpyxl\n"
        f"원본 에러: {e}"
    )

# -----------------------------
# 0) 공통 설정
# -----------------------------
TODAY = pd.Timestamp.today(tz="Asia/Seoul").date()
# 월평균/연속일 판정 등을 위해 충분한 과거치 확보 (약 14개월)
START_DATE = (pd.Timestamp(TODAY) - timedelta(days=420)).strftime("%Y-%m-%d")
END_DATE = pd.Timestamp(TODAY).strftime("%Y-%m-%d")

# 💾 저장 경로 (실행 폴더)
output_path = Path(r"C:/Users/amongpapa/chartup/raw_data") / f"risk_thresholds_{pd.Timestamp(TODAY).strftime('%Y%m%d')}.xlsx"

# -----------------------------
# 1) 티커 맵
# -----------------------------
TICKERS = {
    # KTB / KR rates
    "KR1Y":         "SKTB1YAY Index",
    "KR3Y":         "SKTB3YAY Index",
    "KR10Y":        "SKTB10YY Index",

    # UST & SOFR
    "US10Y":        "USGG10YR Index",
    "TSFR6M":       "TSFR6M Index",
    "TSFR3M":       "TSFR3M Index",
    "TSFR1M":       "TSFR1M Index",
    "SOFR_OIS_1Y":  "USOSFR1 BGN Curncy",   # ✅ 팀장님 지정

    # FX / Equities / Vol
    "USDKRW":       "USDKRW Curncy",
    "KOSPI":        "KOSPI Index",
    "VKOSPI":       "VKOSPI Index",
    "KRW_IV1Y":     "USDKRWV1Y BGN Curncy",

    # KR policy & money
    "KRBASERATE":   "KORP7D Index",         # ✅ 한국 기준금리
    "KRCALL":       "KOCR Index",           # ✅ 시장 콜금리
    "KR_CD3M":      "KWCDC CMPN Curncy",    # ✅ CD 3개월물
    "KR_FIN1Y_AAA": "KRFN1YAA Index",       # 금융채 1년 AAA
    "KR_CORP3Y_AA-":"KRCORP3YAA- Index",    # 회사채 3년 AA-

    # Other markets
    "SPX":          "SPX Index",
    "SX5E":         "SX5E Index",

    # JPY money market
    "JPY_TIBOR3M":  "TI0003M Index",        # ✅ JPY 3M TIBOR

    # FRA-OIS (환경에 따라 틱커 상이할 수 있음)
    "US_FRAOIS_3M": "USSFRAOIS Index",      # ⚠️ 필요시 교체/비활성
}

# 국가별 5Y CDS (D14)
CDS_TICKERS = {
    "Korea":          "KOREA CDS USD SR 5Y D14 Curncy",   # 한국
    "United States":  "US CDS EUR SR 5Y D14 Curncy",
    "Japan":          "JGB CDS USD SR 5Y D14 Curncy",
    "China":          "CHINAGOV CDS USD SR 5Y D14 Curncy",
    "Vietnam":        "VIETNM CDS USD SR 5Y D14 Curncy",
    "Kazakhstan":     "KAZAKS CDS USD SR 5Y D14 Curncy",
    "Germany":        "GERMAN CDS USD SR 5Y D14 Curncy",
    "United Kingdom": "UK CDS USD SR 5Y D14 Curncy",
    "India":          "INDIA CDS USD SR 5Y D14 Curncy",
    "Mexico":         "MEX CDS USD SR 5Y D14 Curncy",
    "Indonesia":      "INDON CDS USD SR 5Y D14 Curncy",
    "Türkiye":        "TURKEY CDS USD SR 5Y D14 Curncy",
    "Canada":         "CANPAC CDS USD SR 5Y D14 Curncy",
    "Hong Kong":      "HONGK CDS USD SR 5Y D14 Curncy",
    "Australia":      "AUSTLA CDS USD SR 5Y D14 Curncy",
    "Philippines":    "PHILIP CDS USD SR 5Y D14 Curncy",
    "Singapore":      "SINGP CDS USD SR 5Y D14 Curncy",    # ✅ 추가 가능
    "UAE":            "DPWDU CDS USD SR 5Y D14 Curncy",    # ✅ 요청사항 반영(기업계열)
}

# -----------------------------
# 2) 필드 우선순위
# -----------------------------
# - 금리: YLD_YTM_MID > YLD_YTM_LAST > PX_LAST
# - 지수/환율/변동성: PX_LAST
# - ★ CDS: LAST_PRICE > MID > PX_LAST  (※ 핵심 수정)
FIELD_PREFS = {}
for k in TICKERS:
    if k in ["KR1Y","KR3Y","KR10Y"]:
        FIELD_PREFS[k] = ["YLD_YTM_MID", "YLD_YTM_LAST", "PX_LAST"]
    elif k in ["US10Y"]:
        FIELD_PREFS[k] = ["PX_LAST", "YLD_YTM_MID"]
    elif k in ["TSFR6M","TSFR3M","TSFR1M","USDKRW","KOSPI","VKOSPI","KRW_IV1Y","SPX","SX5E"]:
        FIELD_PREFS[k] = ["PX_LAST"]
    elif k in ["SOFR_OIS_1Y"]:
        FIELD_PREFS[k] = ["PX_LAST", "YLD_YTM_MID", "LAST_PRICE"]
    elif k in ["KRBASERATE","KRCALL","KR_CD3M","JPY_TIBOR3M","US_FRAOIS_3M"]:
        FIELD_PREFS[k] = ["PX_LAST", "LAST_PRICE"]
    elif k in ["KR_FIN1Y_AAA","KR_CORP3Y_AA-"]:
        FIELD_PREFS[k] = ["PX_LAST", "YLD_YTM_MID", "LAST_PRICE"]
    else:
        FIELD_PREFS[k] = ["PX_LAST"]

# ★ CDS 전용: (이전: PX_MID→PX_LAST) → (변경: LAST_PRICE→MID→PX_LAST)
for name in CDS_TICKERS:
    FIELD_PREFS[name] = ["LAST_PRICE", "MID", "PX_LAST"]  # ★ 수정

# ★ BDH 호출 시 누락 방지를 위해 항상 포함할 Fallback 필드
FALLBACK_FIELDS = {"LAST_PRICE", "MID", "PX_LAST", "PX_MID", "YLD_YTM_MID", "YLD_YTM_LAST"}  # ★ 수정

# -----------------------------
# 3) 임계값
# -----------------------------
THRESHOLDS = {
    "KR3Y_1d_bp": 15,   "KR3Y_10d_bp": 50,
    "KR10Y_1d_bp": 15,  "KR10Y_10d_bp": 45,
    "USFX_1d_pct": 2.0, "USFX_10d_pct": 5.0,
    "KOSPI_1d_down_pct": -3.5, "KOSPI_10d_down_pct": -10.0,
    "VKOSPI_1d_up_pp": 5.0,    "VKOSPI_10d_up_pp": 10.0,
    "KRWIV_1d_pp": 5.0, "KRWIV_10d_pp": 10.0,
    "G3M_dev_bp": 100.0,

    # 추가 임계
    "SPREAD_SOFR1M_vs_OIS1Y_MTD_bp": 150.0,
    "KR_1Y_minus_BASE_5d_level_bp": -24.0,
    "BASE_minus_CALL_bp": 40.0,
    "TSFR3M_prevM_abs_bp": 75.0,
    "JPY_TIBOR3M_prevM_abs_bp": 25.0,
    "KR5YCDS_prevM_bp_3d": 100.0,
    "KR5YCDS_M3ago_bp_3d": 200.0,
    "KR_10Y_3Y_inversion_5d": 5,
    "CDS_prevM_pct_up": 30.0,
    "CorpAAminus_KTB3Y_ratio_prevM_pct": 16.0,
    "Fin1Y_minus_CD3M_MTD_bp": 70.0,
    "Fin1YAAA_minus_KTB1Y_5d_bp": 50.0,
    "SPX_1d_down_pct": -3.0, "SPX_10d_down_pct": -12.0,
    "SX5E_1d_abs_pct": 3.0,  "SX5E_10d_down_pct": -12.0,
    "FRAOIS_prevM_bp": 30.0,
}

# -----------------------------
# 4) 유틸 함수
# -----------------------------
def fetch_hist_with_field_prefs(ticker_map, cds_map, field_prefs, start_date, end_date):
    """
    멀티 필드로 BDH 조회 후, 가용성이 가장 좋은 필드를 채택하여 단일 시계열로 병합
    반환: DataFrame(index=Date, columns=keys)
    """
    all_pairs = list(ticker_map.items()) + list(cds_map.items())
    # ★ 요청 필드 = 우선순위 필드 합집합 + Fallback (누락 방지)
    prefs_fields = {fld for prefs in field_prefs.values() for fld in prefs}
    all_fields = sorted(prefs_fields | FALLBACK_FIELDS)  # ★ 수정

    raw = blp.bdh(
        tickers=[t for _, t in all_pairs],
        flds=all_fields,
        start_date=start_date,
        end_date=end_date,
        Per="D",
        Fill="P",
    )

    # 일부 환경에서 단일 종목/필드 조합일 때 MultiIndex가 아닐 수 있으므로 방어
    if not isinstance(raw.columns, pd.MultiIndex):
        raw.columns = pd.MultiIndex.from_product([[all_pairs[0][1]], all_fields])

    raw.index = pd.to_datetime(raw.index, errors="coerce")
    raw = raw.sort_index()

    out = {}

    # 공통 fallback (일반자산): PX_LAST → LAST_PRICE → PX_MID → MID
    default_general = ["PX_LAST", "LAST_PRICE", "PX_MID", "MID"]
    # 공통 fallback (CDS): LAST_PRICE → MID → PX_LAST
    default_cds = ["LAST_PRICE", "MID", "PX_LAST"]

    # 일반 틱커
    for key, bb_ticker in ticker_map.items():
        prefs = field_prefs.get(key, default_general)  # ★ 수정: 일반도 보강
        ser = None
        for fld in prefs:
            if (bb_ticker, fld) in raw.columns:
                tmp = pd.to_numeric(raw[(bb_ticker, fld)], errors="coerce").ffill().bfill()
                if tmp.notna().sum() > 0:
                    ser = tmp
                    # print(f"[{key}] 사용 필드: {fld}")  # 디버그용
                    break
        out[key] = ser if ser is not None else pd.Series(index=raw.index, dtype=float)

    # CDS(국가명 키)
    for name, bb_ticker in cds_map.items():
        prefs = field_prefs.get(name, default_cds)     # ★ 수정: CDS 기본값 고정
        ser = None
        for fld in prefs:
            if (bb_ticker, fld) in raw.columns:
                tmp = pd.to_numeric(raw[(bb_ticker, fld)], errors="coerce").ffill().bfill()
                if tmp.notna().sum() > 0:
                    ser = tmp
                    # print(f"[CDS:{name}] 사용 필드: {fld}")  # 디버그용
                    break
        out[name] = ser if ser is not None else pd.Series(index=raw.index, dtype=float)

    df = pd.DataFrame(out)
    df.index = pd.to_datetime(df.index, errors="coerce")
    return df

def last_value(series):
    s = series.dropna()
    return float(s.iloc[-1]) if len(s) else np.nan

def bp_change(series, days=1):
    s = series.dropna()
    if len(s) <= days:
        return np.nan
    return float((s.iloc[-1] - s.iloc[-1 - days]) * 100.0)

def pct_change(series, days=1):
    s = series.dropna()
    if len(s) <= days:
        return np.nan
    prev = s.iloc[-1 - days]
    if prev == 0 or np.isnan(prev):
        return np.nan
    return float((s.iloc[-1] / prev - 1.0) * 100.0)

def pp_change(series, days=1):
    s = series.dropna()
    if len(s) <= days:
        return np.nan
    return float(s.iloc[-1] - s.iloc[-1 - days])

def trailing_3m_avg(series):
    """최근 3개월(영업일 약 63개) 평균"""
    s = series.dropna()
    if len(s) < 10:
        return np.nan
    window = min(63, len(s))
    return float(s.iloc[-window:].mean())

def month_avg(series, months_ago=0):
    """
    달력월 평균 (months_ago=0: 당월, 1: 전월, 3: 3개월전)
    - 인덱스가 datetime 변환 가능한지 안전하게 체크
    """
    s = series.dropna()
    if s.empty:
        return np.nan
    dt = pd.to_datetime(s.index, errors="coerce")
    valid = ~pd.isna(dt)
    s, dt = s[valid], dt[valid]
    if len(s) == 0:
        return np.nan
    last_dt = dt.max()
    target = last_dt - pd.DateOffset(months=months_ago)
    mask = (dt.year == target.year) & (dt.month == target.month)
    if not mask.any():
        return np.nan
    # ★ boolean mask 인덱싱 안정화
    return float(s[mask].mean())  # ★ 수정

def consec_last_n(bool_series, n):
    """마지막 N영업일 모두 True 인지"""
    s = bool_series.dropna().astype(bool)
    if len(s) < n:
        return False
    return bool(s.iloc[-n:].all())

def has_data(series, min_points=5):
    """데이터 유효성 체크: 결측 제거 후 최소 개수 확보"""
    return series.dropna().shape[0] >= min_points

# -----------------------------
# 5) 데이터 수집
# -----------------------------
hist = fetch_hist_with_field_prefs(
    ticker_map=TICKERS,
    cds_map=CDS_TICKERS,
    field_prefs=FIELD_PREFS,
    start_date=START_DATE,
    end_date=END_DATE,
)

# series 접근용 맵 구성 (키: TICKERS/국가명)
series_map = {}
for k in TICKERS.keys():
    if k in hist.columns:
        series_map[k] = hist[k]
for name in CDS_TICKERS.keys():
    if name in hist.columns:
        series_map[name] = hist[name]

# -----------------------------
# 6) 임계 로직
# -----------------------------
rows = []

# ---- (A) 원화금리 - 국고 3년 ----
if has_data(series_map.get("KR3Y", pd.Series(dtype=float))):
    kr3y_1d = bp_change(series_map["KR3Y"], 1)
    kr3y_10d = bp_change(series_map["KR3Y"], 10)
    rows.append({
        "metric": "KR 3Y KTB Yield",
        "ticker": TICKERS["KR3Y"],
        "latest": last_value(series_map["KR3Y"]),
        "chg_1d": f"{kr3y_1d:.1f}bp" if pd.notna(kr3y_1d) else np.nan,
        "threshold_1d": "±15bp",
        "breach_1d": (abs(kr3y_1d) >= THRESHOLDS["KR3Y_1d_bp"]) if pd.notna(kr3y_1d) else np.nan,
        "chg_10d": f"{kr3y_10d:.1f}bp" if pd.notna(kr3y_10d) else np.nan,
        "threshold_10d": "±50bp",
        "breach_10d": (abs(kr3y_10d) >= THRESHOLDS["KR3Y_10d_bp"]) if pd.notna(kr3y_10d) else np.nan,
        "note": "원화 3Y: 수익률 bp 기준",
    })

# ---- (B) 원화금리 - 국고 10년 ----
if has_data(series_map.get("KR10Y", pd.Series(dtype=float))):
    kr10y_1d = bp_change(series_map["KR10Y"], 1)
    kr10y_10d = bp_change(series_map["KR10Y"], 10)
    rows.append({
        "metric": "KR 10Y KTB Yield",
        "ticker": TICKERS["KR10Y"],
        "latest": last_value(series_map["KR10Y"]),
        "chg_1d": f"{kr10y_1d:.1f}bp" if pd.notna(kr10y_1d) else np.nan,
        "threshold_1d": "±15bp",
        "breach_1d": (abs(kr10y_1d) >= THRESHOLDS["KR10Y_1d_bp"]) if pd.notna(kr10y_1d) else np.nan,
        "chg_10d": f"{kr10y_10d:.1f}bp" if pd.notna(kr10y_10d) else np.nan,
        "threshold_10d": "±45bp",
        "breach_10d": (abs(kr10y_10d) >= THRESHOLDS["KR10Y_10d_bp"]) if pd.notna(kr10y_10d) else np.nan,
        "note": "원화 10Y: 수익률 bp 기준",
    })

# ---- (C) 미10Y: 3M 평균 대비 ±100bp ----
if has_data(series_map.get("US10Y", pd.Series(dtype=float))):
    us10y_last = last_value(series_map["US10Y"])
    us10y_3m = trailing_3m_avg(series_map["US10Y"])
    dev_bp = (us10y_last - us10y_3m) * 100.0 if pd.notna(us10y_last) and pd.notna(us10y_3m) else np.nan
    rows.append({
        "metric": "US 10Y vs 3M Avg",
        "ticker": TICKERS["US10Y"],
        "latest": us10y_last,
        "breach_3m": (abs(dev_bp) >= THRESHOLDS["G3M_dev_bp"]) if pd.notna(dev_bp) else np.nan,
        "note": f"3M avg={us10y_3m:.4f}, dev={dev_bp:.1f}bp; 임계±{THRESHOLDS['G3M_dev_bp']}bp" if pd.notna(dev_bp) else "데이터 부족",
    })

# ---- (D) TSFR 6M: 3M 평균 대비 ±100bp ----
if has_data(series_map.get("TSFR6M", pd.Series(dtype=float))):
    ts6_last = last_value(series_map["TSFR6M"])
    ts6_3m = trailing_3m_avg(series_map["TSFR6M"])
    dev_bp = (ts6_last - ts6_3m) * 100.0 if pd.notna(ts6_last) and pd.notna(ts6_3m) else np.nan
    rows.append({
        "metric": "TSFR 6M vs 3M Avg",
        "ticker": TICKERS["TSFR6M"],
        "latest": ts6_last,
        "breach_3m": (abs(dev_bp) >= THRESHOLDS["G3M_dev_bp"]) if pd.notna(dev_bp) else np.nan,
        "note": f"3M avg={ts6_3m:.4f}, dev={dev_bp:.1f}bp; 임계±{THRESHOLDS['G3M_dev_bp']}bp" if pd.notna(dev_bp) else "데이터 부족",
    })

# ---- (E) USDKRW: 1일 ±2%, 10일 ±5% ----
if has_data(series_map.get("USDKRW", pd.Series(dtype=float))):
    krw_1d = pct_change(series_map["USDKRW"], 1)
    krw_10d = pct_change(series_map["USDKRW"], 10)
    rows.append({
        "metric": "USDKRW Spot",
        "ticker": TICKERS["USDKRW"],
        "latest": last_value(series_map["USDKRW"]),
        "chg_1d": f"{krw_1d:.2f}%" if pd.notna(krw_1d) else np.nan,
        "threshold_1d": "±2.0%",
        "breach_1d": (abs(krw_1d) >= THRESHOLDS["USFX_1d_pct"]) if pd.notna(krw_1d) else np.nan,
        "chg_10d": f"{krw_10d:.2f}%" if pd.notna(krw_10d) else np.nan,
        "threshold_10d": "±5.0%",
        "breach_10d": (abs(krw_10d) >= THRESHOLDS["USFX_10d_pct"]) if pd.notna(krw_10d) else np.nan,
        "note": "원/달러 환율: % 기준",
    })

# ---- (F) KOSPI: 1일 -3.5%, 10일 -10% (하락만) ----
if has_data(series_map.get("KOSPI", pd.Series(dtype=float))):
    k1 = pct_change(series_map["KOSPI"], 1)
    k10 = pct_change(series_map["KOSPI"], 10)
    rows.append({
        "metric": "KOSPI Index",
        "ticker": TICKERS["KOSPI"],
        "latest": last_value(series_map["KOSPI"]),
        "chg_1d": f"{k1:.2f}%" if pd.notna(k1) else np.nan,
        "threshold_1d": "≤ -3.5%",
        "breach_1d": (k1 <= THRESHOLDS["KOSPI_1d_down_pct"]) if pd.notna(k1) else np.nan,
        "chg_10d": f"{k10:.2f}%" if pd.notna(k10) else np.nan,
        "threshold_10d": "≤ -10.0%",
        "breach_10d": (k10 <= THRESHOLDS["KOSPI_10d_down_pct"]) if pd.notna(k10) else np.nan,
        "note": "하락 방향만 트리거",
    })

# ---- (G) VKOSPI: 1일 +5pp, 10일 +10pp (상승만) ----
if has_data(series_map.get("VKOSPI", pd.Series(dtype=float))):
    v1 = pp_change(series_map["VKOSPI"], 1)
    v10 = pp_change(series_map["VKOSPI"], 10)
    rows.append({
        "metric": "VKOSPI (Vol Index)",
        "ticker": TICKERS["VKOSPI"],
        "latest": last_value(series_map["VKOSPI"]),
        "chg_1d": f"{v1:.2f}pp" if pd.notna(v1) else np.nan,
        "threshold_1d": "≥ +5.0pp",
        "breach_1d": (v1 >= THRESHOLDS["VKOSPI_1d_up_pp"]) if pd.notna(v1) else np.nan,
        "chg_10d": f"{v10:.2f}pp" if pd.notna(v10) else np.nan,
        "threshold_10d": "≥ +10.0pp",
        "breach_10d": (v10 >= THRESHOLDS["VKOSPI_10d_up_pp"]) if pd.notna(v10) else np.nan,
        "note": "상승만 트리거(pp)",
    })

# ---- (H) USDKRW 1Y IV: 1일 ±5pp, 10일 ±10pp ----
if has_data(series_map.get("KRW_IV1Y", pd.Series(dtype=float))):
    iv1 = pp_change(series_map["KRW_IV1Y"], 1)
    iv10 = pp_change(series_map["KRW_IV1Y"], 10)
    rows.append({
        "metric": "USDKRW 1Y Implied Vol",
        "ticker": TICKERS["KRW_IV1Y"],
        "latest": last_value(series_map["KRW_IV1Y"]),
        "chg_1d": f"{iv1:.2f}pp" if pd.notna(iv1) else np.nan,
        "threshold_1d": "±5.0pp",
        "breach_1d": (abs(iv1) >= THRESHOLDS["KRWIV_1d_pp"]) if pd.notna(iv1) else np.nan,
        "chg_10d": f"{iv10:.2f}pp" if pd.notna(iv10) else np.nan,
        "threshold_10d": "±10.0pp",
        "breach_10d": (abs(iv10) >= THRESHOLDS["KRWIV_10d_pp"]) if pd.notna(iv10) else np.nan,
        "note": "절대 pp 기준",
    })

# ---- (I) 외화 월평균 장단기: (SOFR OIS 1Y - TSFR 1M) MTD ≥ +150bp ----
if has_data(series_map.get("SOFR_OIS_1Y", pd.Series(dtype=float))) and has_data(series_map.get("TSFR1M", pd.Series(dtype=float))):
    ois1y_mtd = month_avg(series_map["SOFR_OIS_1Y"], 0)
    tsfr1m_mtd = month_avg(series_map["TSFR1M"], 0)
    mtd_spread = (ois1y_mtd - tsfr1m_mtd) * 100.0 if pd.notna(ois1y_mtd) and pd.notna(tsfr1m_mtd) else np.nan
    rows.append({
        "metric": "USD OIS 1Y - TSFR 1M (MTD avg)",
        "ticker": f"{TICKERS['SOFR_OIS_1Y']} vs {TICKERS['TSFR1M']}",
        "chg_1d": f"{mtd_spread:.1f}bp (MTD spread)" if pd.notna(mtd_spread) else np.nan,
        "threshold_1d": f"≥ +{THRESHOLDS['SPREAD_SOFR1M_vs_OIS1Y_MTD_bp']:.0f}bp",
        "breach_1d": (mtd_spread >= THRESHOLDS["SPREAD_SOFR1M_vs_OIS1Y_MTD_bp"]) if pd.notna(mtd_spread) else np.nan,
        "note": f"MTD OIS1Y={ois1y_mtd:.4f}, TSFR1M={tsfr1m_mtd:.4f}" if pd.notna(mtd_spread) else "데이터/틱커 확인 필요",
    })

# ---- (J) KR 1Y - 기준금리: 5영업일 연속 < -24bp ----
if has_data(series_map.get("KR1Y", pd.Series(dtype=float))) and has_data(series_map.get("KRBASERATE", pd.Series(dtype=float))):
    spr_bp = (series_map["KR1Y"] - series_map["KRBASERATE"]) * 100.0
    rows.append({
        "metric": "KR 1Y - BaseRate (level)",
        "ticker": f"{TICKERS['KR1Y']} - {TICKERS['KRBASERATE']}",
        "latest": float(spr_bp.dropna().iloc[-1]) if spr_bp.dropna().size else np.nan,
        "threshold_1d": "5영업일 연속 < -24bp",
        "breach_1d": consec_last_n(spr_bp < THRESHOLDS["KR_1Y_minus_BASE_5d_level_bp"], 5) if spr_bp.dropna().size else np.nan,
        "note": "레벨 기준(일별 스프레드)",
    })

# ---- (K) 기준금리 - 콜금리: > +40bp ----
if has_data(series_map.get("KRBASERATE", pd.Series(dtype=float))) and has_data(series_map.get("KRCALL", pd.Series(dtype=float))):
    base_call = (series_map["KRBASERATE"] - series_map["KRCALL"]) * 100.0
    rows.append({
        "metric": "KR Base - Call (level)",
        "ticker": f"{TICKERS['KRBASERATE']} - {TICKERS['KRCALL']}",
        "latest": float(base_call.dropna().iloc[-1]) if base_call.dropna().size else np.nan,
        "threshold_1d": f"> +{THRESHOLDS['BASE_minus_CALL_bp']:.0f}bp",
        "breach_1d": (base_call.dropna().iloc[-1] > THRESHOLDS["BASE_minus_CALL_bp"]) if base_call.dropna().size else np.nan,
        "note": "레벨 기준",
    })

# ---- (L) TSFR 3M: MTD-PrevM 절대변화 > 75bp ----
if has_data(series_map.get("TSFR3M", pd.Series(dtype=float))):
    ts3_cur = month_avg(series_map["TSFR3M"], 0)
    ts3_prev = month_avg(series_map["TSFR3M"], 1)
    ts3_diff = (ts3_cur - ts3_prev) * 100.0 if pd.notna(ts3_cur) and pd.notna(ts3_prev) else np.nan
    rows.append({
        "metric": "TSFR 3M (MTD - PrevM)",
        "ticker": TICKERS["TSFR3M"],
        "chg_1d": f"{ts3_diff:.1f}bp (Δavg)" if pd.notna(ts3_diff) else np.nan,
        "threshold_1d": f"abs(Δ) > {THRESHOLDS['TSFR3M_prevM_abs_bp']:.0f}bp",
        "breach_1d": (abs(ts3_diff) > THRESHOLDS["TSFR3M_prevM_abs_bp"]) if pd.notna(ts3_diff) else np.nan,
        "note": f"MTD={ts3_cur:.4f}, PrevM={ts3_prev:.4f}" if pd.notna(ts3_diff) else "데이터/틱커 확인 필요",
    })

# ---- (M) JPY 3M TIBOR: MTD-PrevM 절대변화 > 25bp ----
if has_data(series_map.get("JPY_TIBOR3M", pd.Series(dtype=float))):
    tib_cur = month_avg(series_map["JPY_TIBOR3M"], 0)
    tib_prev = month_avg(series_map["JPY_TIBOR3M"], 1)
    tib_diff = (tib_cur - tib_prev) * 100.0 if pd.notna(tib_cur) and pd.notna(tib_prev) else np.nan
    rows.append({
        "metric": "JPY TIBOR 3M (MTD - PrevM)",
        "ticker": TICKERS["JPY_TIBOR3M"],
        "chg_1d": f"{tib_diff:.1f}bp (Δavg)" if pd.notna(tib_diff) else np.nan,
        "threshold_1d": f"abs(Δ) > {THRESHOLDS['JPY_TIBOR3M_prevM_abs_bp']:.0f}bp",
        "breach_1d": (abs(tib_diff) > THRESHOLDS["JPY_TIBOR3M_prevM_abs_bp"]) if pd.notna(tib_diff) else np.nan,
        "note": f"MTD={tib_cur:.4f}, PrevM={tib_prev:.4f}" if pd.notna(tib_diff) else "데이터/틱커 확인 필요",
    })

# ---- (N) 한국 5Y CDS: PrevM +100bp 3D / M-3 +200bp 3D ----
if has_data(series_map.get("Korea", pd.Series(dtype=float))):
    cds_kr = series_map["Korea"]
    prevM_avg = month_avg(cds_kr, 1)
    m3_avg = month_avg(cds_kr, 3)

    if pd.notna(prevM_avg):
        dev_prev = cds_kr - prevM_avg
        rows.append({
            "metric": "KR 5Y CDS vs PrevM (3D consec)",
            "ticker": CDS_TICKERS["Korea"],
            "latest": last_value(cds_kr),
            "threshold_1d": f"> +{THRESHOLDS['KR5YCDS_prevM_bp_3d']:.0f}bp for 3D",
            "breach_1d": consec_last_n(dev_prev > THRESHOLDS["KR5YCDS_prevM_bp_3d"], 3),
            "note": f"PrevM avg={prevM_avg:.1f}bp",
        })

    if pd.notna(m3_avg):
        dev_m3 = cds_kr - m3_avg
        rows.append({
            "metric": "KR 5Y CDS vs M-3 (3D consec)",
            "ticker": CDS_TICKERS["Korea"],
            "latest": last_value(cds_kr),
            "threshold_1d": f"> +{THRESHOLDS['KR5YCDS_M3ago_bp_3d']:.0f}bp for 3D",
            "breach_1d": consec_last_n(dev_m3 > THRESHOLDS["KR5YCDS_M3ago_bp_3d"], 3),
            "note": f"M-3 avg={m3_avg:.1f}bp",
        })

# ---- (O) KR Term Spread (10Y-3Y): 5D 역전 지속 ----
if has_data(series_map.get("KR10Y", pd.Series(dtype=float))) and has_data(series_map.get("KR3Y", pd.Series(dtype=float))):
    term_spread = (series_map["KR10Y"] - series_map["KR3Y"]) * 100.0
    rows.append({
        "metric": "KR Term Spread 10Y-3Y (5D inversion)",
        "ticker": f"{TICKERS['KR10Y']} - {TICKERS['KR3Y']}",
        "latest": float(term_spread.dropna().iloc[-1]) if term_spread.dropna().size else np.nan,
        "threshold_1d": "< 0bp for 5D",
        "breach_1d": consec_last_n(term_spread <= 0.0, THRESHOLDS["KR_10Y_3Y_inversion_5d"]) if term_spread.dropna().size else np.nan,
        "note": "10Y-3Y ≤ 0bp 상태 5D 연속",
    })

# ---- (P) 국가별 CDS 17개국: 전월 평균 대비 +30% 상승 ----
for country, bb in CDS_TICKERS.items():
    if country == "Korea":
        continue  # 한국은 위에서 bp 기준 3D 연속 로직 적용
    s = series_map.get(country, pd.Series(dtype=float))
    if has_data(s):
        mtd, prev = month_avg(s, 0), month_avg(s, 1)
        pct_up = ((mtd / prev - 1.0) * 100.0) if (pd.notna(mtd) and pd.notna(prev) and prev != 0) else np.nan
        rows.append({
            "metric": f"CDS 5Y: {country} (MTD vs PrevM)",
            "ticker": bb,
            "latest": last_value(s),
            "chg_1d": f"{pct_up:.1f}%" if pd.notna(pct_up) else np.nan,
            "threshold_1d": f"> +{THRESHOLDS['CDS_prevM_pct_up']:.0f}%",
            "breach_1d": (pct_up > THRESHOLDS["CDS_prevM_pct_up"]) if pd.notna(pct_up) else np.nan,
            "note": f"MTD={mtd:.1f}, PrevM={prev:.1f}" if pd.notna(pct_up) else "데이터/틱커 확인 필요",
        })

# ---- (Q) (회사채/국고) 3Y 비율: 전월평균 대비 +16% 상승 ----
if has_data(series_map.get("KR3Y", pd.Series(dtype=float))) and has_data(series_map.get("KR_CORP3Y_AA-", pd.Series(dtype=float))):
    ktb3y_mtd  = month_avg(series_map["KR3Y"], 0)
    corp3y_mtd = month_avg(series_map["KR_CORP3Y_AA-"], 0)
    ktb3y_prev  = month_avg(series_map["KR3Y"], 1)
    corp3y_prev = month_avg(series_map["KR_CORP3Y_AA-"], 1)

    ratio_cur  = (ktb3y_mtd / corp3y_mtd) if (pd.notna(ktb3y_mtd) and pd.notna(corp3y_mtd) and corp3y_mtd != 0) else np.nan
    ratio_prev = (ktb3y_prev / corp3y_prev) if (pd.notna(ktb3y_prev) and pd.notna(corp3y_prev) and corp3y_prev != 0) else np.nan
    ratio_pct  = ((ratio_cur / ratio_prev - 1.0) * 100.0) if (pd.notna(ratio_cur) and pd.notna(ratio_prev) and ratio_prev != 0) else np.nan

    rows.append({
        "metric": "KTB3Y / Corp(AA-) 3Y (MTD vs PrevM)",
        "ticker": "KR3Y / KR_CORP3Y_AA-",
        "chg_1d": f"{ratio_pct:.1f}%" if pd.notna(ratio_pct) else np.nan,
        "threshold_1d": f"> +{THRESHOLDS['CorpAAminus_KTB3Y_ratio_prevM_pct']:.0f}%",
        "breach_1d": (ratio_pct > THRESHOLDS["CorpAAminus_KTB3Y_ratio_prevM_pct"]) if pd.notna(ratio_pct) else np.nan,
        "note": f"MTD={ratio_cur:.4f}, PrevM={ratio_prev:.4f}" if pd.notna(ratio_pct) else "데이터/틱커 확인 필요",
    })

# ---- (R) 월평균 장단기 (금융채1Y - CD3M): MTD 스프레드 ≥ +70bp ----
if has_data(series_map.get("KR_FIN1Y_AAA", pd.Series(dtype=float))) and has_data(series_map.get("KR_CD3M", pd.Series(dtype=float))):
    fin1y_mtd = month_avg(series_map["KR_FIN1Y_AAA"], 0)
    cd3m_mtd  = month_avg(series_map["KR_CD3M"], 0)
    fin_cd_bp = (fin1y_mtd - cd3m_mtd) * 100.0 if pd.notna(fin1y_mtd) and pd.notna(cd3m_mtd) else np.nan
    rows.append({
        "metric": "(MTD) Fin 1Y - CD 3M",
        "ticker": f"{TICKERS['KR_FIN1Y_AAA']} - {TICKERS['KR_CD3M']}",
        "chg_1d": f"{fin_cd_bp:.1f}bp" if pd.notna(fin_cd_bp) else np.nan,
        "threshold_1d": f"≥ +{THRESHOLDS['Fin1Y_minus_CD3M_MTD_bp']:.0f}bp",
        "breach_1d": (fin_cd_bp >= THRESHOLDS["Fin1Y_minus_CD3M_MTD_bp"]) if pd.notna(fin_cd_bp) else np.nan,
        "note": f"MTD Fin1Y={fin1y_mtd:.4f}, CD3M={cd3m_mtd:.4f}" if pd.notna(fin_cd_bp) else "데이터/틱커 확인 필요",
    })

# ---- (S) Fin1Y AAA - KTB1Y: 5영업일 연속 ≥ +50bp ----
if has_data(series_map.get("KR_FIN1Y_AAA", pd.Series(dtype=float))) and has_data(series_map.get("KR1Y", pd.Series(dtype=float))):
    fin_minus_ktb1y = (series_map["KR_FIN1Y_AAA"] - series_map["KR1Y"]) * 100.0
    rows.append({
        "metric": "Fin 1Y(AAA) - KTB 1Y (5D consec ≥50bp)",
        "ticker": f"{TICKERS['KR_FIN1Y_AAA']} - {TICKERS['KR1Y']}",
        "latest": float(fin_minus_ktb1y.dropna().iloc[-1]) if fin_minus_ktb1y.dropna().size else np.nan,
        "threshold_1d": f"≥ +{THRESHOLDS['Fin1YAAA_minus_KTB1Y_5d_bp']:.0f}bp for 5D",
        "breach_1d": consec_last_n(fin_minus_ktb1y >= THRESHOLDS["Fin1YAAA_minus_KTB1Y_5d_bp"], 5) if fin_minus_ktb1y.dropna().size else np.nan,
        "note": "레벨 기준(일별 스프레드)",
    })

# ---- (T) S&P 500: 1일 ≤ -3%, 10일 ≤ -12% ----
if has_data(series_map.get("SPX", pd.Series(dtype=float))):
    spx_1d = pct_change(series_map["SPX"], 1)
    spx_10d = pct_change(series_map["SPX"], 10)
    rows.append({
        "metric": "S&P 500",
        "ticker": TICKERS["SPX"],
        "latest": last_value(series_map["SPX"]),
        "chg_1d": f"{spx_1d:.2f}%" if pd.notna(spx_1d) else np.nan,
        "threshold_1d": "≤ -3.0%",
        "breach_1d": (spx_1d <= THRESHOLDS["SPX_1d_down_pct"]) if pd.notna(spx_1d) else np.nan,
        "chg_10d": f"{spx_10d:.2f}%" if pd.notna(spx_10d) else np.nan,
        "threshold_10d": "≤ -12.0%",
        "breach_10d": (spx_10d <= THRESHOLDS["SPX_10d_down_pct"]) if pd.notna(spx_10d) else np.nan,
        "note": "하락만 트리거",
    })

# ---- (U) EuroStoxx50: 1일 |Δ| ≥ 3%, 10일 ≤ -12% ----
if has_data(series_map.get("SX5E", pd.Series(dtype=float))):
    sx_1d = pct_change(series_map["SX5E"], 1)
    sx_10d = pct_change(series_map["SX5E"], 10)
    rows.append({
        "metric": "EuroStoxx50",
        "ticker": TICKERS["SX5E"],
        "latest": last_value(series_map["SX5E"]),
        "chg_1d": f"{sx_1d:.2f}%" if pd.notna(sx_1d) else np.nan,
        "threshold_1d": f"abs ≥ {THRESHOLDS['SX5E_1d_abs_pct']:.1f}%",
        "breach_1d": (abs(sx_1d) >= THRESHOLDS["SX5E_1d_abs_pct"]) if pd.notna(sx_1d) else np.nan,
        "chg_10d": f"{sx_10d:.2f}%" if pd.notna(sx_10d) else np.nan,
        "threshold_10d": "≤ -12.0%",
        "breach_10d": (sx_10d <= THRESHOLDS["SX5E_10d_down_pct"]) if pd.notna(sx_10d) else np.nan,
        "note": "1D는 절대값, 10D는 하락만",
    })

# ---- (V) 3M FRA-OIS: PrevM 대비 +30bp ----
if has_data(series_map.get("US_FRAOIS_3M", pd.Series(dtype=float))):
    fraois_mtd  = month_avg(series_map["US_FRAOIS_3M"], 0)
    fraois_prev = month_avg(series_map["US_FRAOIS_3M"], 1)
    diff_bp = (fraois_mtd - fraois_prev) * 100.0 if pd.notna(fraois_mtd) and pd.notna(fraois_prev) else np.nan
    rows.append({
        "metric": "USD 3M FRA-OIS (MTD - PrevM)",
        "ticker": TICKERS["US_FRAOIS_3M"],
        "chg_1d": f"{diff_bp:.1f}bp (Δavg)" if pd.notna(diff_bp) else np.nan,
        "threshold_1d": f"> +{THRESHOLDS['FRAOIS_prevM_bp']:.0f}bp",
        "breach_1d": (diff_bp > THRESHOLDS["FRAOIS_prevM_bp"]) if pd.notna(diff_bp) else np.nan,
        "note": f"MTD={fraois_mtd:.2f}, PrevM={fraois_prev:.2f}" if pd.notna(diff_bp) else "데이터/틱커 확인 필요",
    })

# -----------------------------
# 7) 엑셀 저장 (alerts + raw_data)
# -----------------------------
alerts_df = pd.DataFrame(rows)

order_cols = [
    "metric","ticker","latest",
    "chg_1d","threshold_1d","breach_1d",
    "chg_10d","threshold_10d","breach_10d",
    "breach_3m","note",
]
for c in order_cols:
    if c not in alerts_df.columns:
        alerts_df[c] = np.nan

alerts_df = alerts_df[order_cols]

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    alerts_df.to_excel(writer, sheet_name="alerts", index=False)
    raw_df = hist.copy()
    raw_df.index.name = "Date"
    raw_df.reset_index().to_excel(writer, sheet_name="raw_data", index=False)

print(f"✅ 저장 완료: {output_path}")
print("▶ alerts 미리보기 (상위 20행):")
print(alerts_df.head(20))


#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
리스크 임계치 초과 내역 텔레그램 전송 스크립트

목적:
- 기존 'risk_thresholds_YYYYMMDD.xlsx' (alerts 시트)를 읽어서
- breach_1d / breach_10d / breach_3m 중 하나라도 True인 항목만 골라
- 텔레그램으로 요약 메시지를 전송하고, 마지막 줄에 대시보드 링크를 추가한다.

전제:
- 앞서 팀장님이 실행하신 Bloomberg/xbbg 스크립트와 동일한 날짜 기준으로
  risk_thresholds_YYYYMMDD.xlsx 파일이 생성되어 있어야 한다.
- 텔레그램 봇 토큰과 채널/채팅 ID 가 준비되어 있어야 한다.

필요 패키지:
    pip install pandas requests openpyxl
"""

from pathlib import Path
from datetime import date
import pandas as pd
import numpy as np
import requests

# ==========================
# 1. 텔레그램 설정값
# ==========================
BOT_TOKEN = "8432426313:AAEdvkoFEozZE-1F0ARc82i_JXCWA4fPfgE"  # 팀장님 봇 토큰
CHAT_ID   = "-4956067497"                                     # 팀장님 채널/채팅 ID

TELEGRAM_API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"

# ==========================
# 2. 파일 경로 설정
# ==========================
# - risk_thresholds_YYYYMMDD.xlsx 가 저장된 폴더
OUTPUT_DIR = Path(r"C:/Users/amongpapa/chartup/raw_data")

# - 기본은 "오늘 날짜" 파일을 찾도록 구성
#   (필요하면 main에서 execution_date를 인자로 받도록 확장해도 됨)
def get_risk_excel_path(target_date: date | None = None) -> Path:
    """
    대상 날짜의 risk_thresholds_YYYYMMDD.xlsx 경로를 반환하는 헬퍼 함수.
    target_date를 넘기지 않으면 오늘 날짜 기준으로 파일명을 만든다.
    """
    if target_date is None:
        target_date = date.today()
    fname = f"risk_thresholds_{target_date:%Y%m%d}.xlsx"
    return OUTPUT_DIR / fname


# ==========================
# 3. 엑셀에서 breach 항목 추출
# ==========================
def load_breach_rows(excel_path: Path) -> pd.DataFrame:
    """
    risk_thresholds 엑셀 파일에서 alerts 시트를 읽고,
    breach_1d / breach_10d / breach_3m 중 하나라도 True인 행만 필터링해서 반환.

    반환:
        breach_df (DataFrame): breach 행만 모은 요약 테이블
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"임계치 파일을 찾을 수 없습니다: {excel_path}")

    # alerts 시트 읽기
    df = pd.read_excel(excel_path, sheet_name="alerts")

    # breach 컬럼들만 자동 탐색 (breach로 시작하는 모든 컬럼)
    breach_cols = [c for c in df.columns if c.startswith("breach")]
    if not breach_cols:
        # breach 컬럼이 아예 없으면 빈 DF 반환
        return df.iloc[0:0].copy()

    # NaN → False 로 채우고 행 단위로 OR 조건
    mask = df[breach_cols].fillna(False).any(axis=1)
    breach_df = df.loc[mask].copy()

    return breach_df


# ==========================
# 4. 텔레그램 메시지 텍스트 구성
# ==========================
def format_value(x) -> str:
    """
    NaN, None 은 빈 문자열로 처리하고, 나머지는 문자열로 변환하는 단순 헬퍼.
    """
    if pd.isna(x):
        return ""
    return str(x)


def build_message_from_breach_df(breach_df: pd.DataFrame, target_date: date) -> str:
    """
    breach_df를 기반으로 텔레그램에 보낼 메시지 본문을 구성한다.

    규칙:
    - 헤더: 날짜 + 제목
    - 각 항목:
        * metric (지표 이름)
        * (필요 시) latest 값
        * breach_1d / breach_10d / breach_3m 중 True인 것만 골라서
          변화량 + 기준을 한 줄씩 bullet 로 표시
        * note가 있으면 마지막에 괄호로 추가
    - 마지막 줄에 risk_monitor 링크 추가
    """
    header_lines = [
        f"[리스크 임계치 초과 알림]",
        f"{target_date:%Y-%m-%d} 기준",
        "",
    ]

    body_lines: list[str] = []

    if breach_df.empty:
        # 임계치 초과 항목이 없을 때
        body_lines.append("오늘은 설정된 임계수준을 초과한 지표가 없습니다.")
    else:
        # 각 행별로 텍스트 정리
        for idx, row in breach_df.iterrows():
            metric = format_value(row.get("metric", ""))
            ticker = format_value(row.get("ticker", ""))
            latest = row.get("latest", np.nan)
            latest_str = format_value(latest)

            line_header = f"• {metric}"
            if ticker:
                line_header += f" ({ticker})"
            body_lines.append(line_header)

            # 어떤 breach가 발생했는지 정리
            # 1일
            if bool(row.get("breach_1d", False)):
                chg_1d = format_value(row.get("chg_1d", ""))
                th_1d  = format_value(row.get("threshold_1d", ""))
                body_lines.append(f"   - 1일 변화: {chg_1d} (기준 {th_1d})")

            # 10일
            if bool(row.get("breach_10d", False)):
                chg_10d = format_value(row.get("chg_10d", ""))
                th_10d  = format_value(row.get("threshold_10d", ""))
                body_lines.append(f"   - 10일 변화: {chg_10d} (기준 {th_10d})")

            # 3개월/기타 기준 (breach_3m 컬럼)
            if bool(row.get("breach_3m", False)):
                note = format_value(row.get("note", ""))
                if note:
                    body_lines.append(f"   - 3개월/평균 기준 초과: {note}")
                else:
                    body_lines.append(f"   - 3개월/평균 기준 초과")

            # latest 값이 의미 있을 경우 한 줄 추가 (원치 않으면 주석 처리 가능)
            if latest_str:
                body_lines.append(f"   - 현재 수준: {latest_str}")

            # 항목 간 구분용 빈 줄
            body_lines.append("")

    # footer: 링크 추가
    footer_lines = [
        "",
        "자세한 지표는 내부 대시보드에서 확인해주세요.",
        "https://chartupndown.com/risk_monitor",
    ]

    # 전체 메시지 합치기
    full_message = "\n".join(header_lines + body_lines + footer_lines)
    return full_message


# ==========================
# 5. 텔레그램 전송 함수
# ==========================
def send_telegram_message(text: str) -> None:
    """
    텔레그램 sendMessage API를 호출하여 지정된 CHAT_ID로 메시지를 전송한다.
    """
    payload = {
        "chat_id": CHAT_ID,
        "text": text,
        "parse_mode": "HTML",  # 필요시 Markdown 으로 변경 가능
    }
    try:
        resp = requests.post(TELEGRAM_API_URL, data=payload, timeout=10)
        resp.raise_for_status()
        print("✅ 텔레그램 전송 성공")
    except requests.RequestException as e:
        print("⚠️ 텔레그램 전송 중 오류 발생:", e)


# ==========================
# 6. 메인 실행 흐름
# ==========================
def send_risk_alert_via_telegram(target_date: date | None = None) -> None:
    """
    통합 실행 함수:
    1) 대상 날짜의 risk_thresholds 엑셀 경로 계산
    2) breach 행만 필터링
    3) 텍스트 메시지 구성
    4) 텔레그램 전송
    """
    if target_date is None:
        target_date = date.today()

    excel_path = get_risk_excel_path(target_date)
    print(f"▶ 임계치 파일: {excel_path}")

    breach_df = load_breach_rows(excel_path)
    print(f"▶ 임계치 초과 지표 수: {len(breach_df)}")

    msg = build_message_from_breach_df(breach_df, target_date)
    print("▶ 전송 메시지 미리보기:")
    print("=" * 60)
    print(msg)
    print("=" * 60)

    # 실제 텔레그램 전송
    send_telegram_message(msg)


# ==========================
# 7. 직접 실행 시 진입점
# ==========================
if __name__ == "__main__":
    # 기본은 오늘 날짜 기준으로 동작
    send_risk_alert_via_telegram()

    # 특정 날짜 파일을 보내고 싶으면 예시처럼 호출
    # from datetime import datetime
    # send_risk_alert_via_telegram(datetime(2025, 11, 17).date())


# ───────────────────────────────────────────────────────────────────────────────
# ▶ 필요 패키지 설치 (한 번만 실행)
# pip install xbbg blpapi pandas openpyxl
# ───────────────────────────────────────────────────────────────────────────────

import os
from datetime import datetime, timedelta
import pandas as pd
import blpapi                                   # Bloomberg low‑level API
from xbbg import blp                            # xbbg 래퍼

# 1) 환경 설정
# ───────────────────────────────────────────────────────────────────────────────
# 입력 파일 경로: indicator.xlsx
input_path = r"C:\Users\amongpapa\chartup\go_scen\data\indicator.xlsx"

# 출력을 저장할 폴더 (set)
output_dir = r"C:\Users\amongpapa\chartup\go_scen\data\set"

# 조회 기간: 오늘 기준 과거 1년
today      = datetime.today()
end_date   = today.strftime("%Y-%m-%d")                        # 예: '2025-07-29'
start_date = (today - timedelta(days=365)).strftime("%Y-%m-%d")  # 예: '2024-07-29'

# 2) 출력 폴더 준비
# ───────────────────────────────────────────────────────────────────────────────
os.makedirs(output_dir, exist_ok=True)  # 없으면 생성

# 3) indicator.xlsx 읽기
# ───────────────────────────────────────────────────────────────────────────────
df_ind = pd.read_excel(input_path, dtype=str)  # 모든 칼럼을 문자열로 로드

# 필수 칼럼 확인
required_cols = {'Indicator_ID', 'Bloomberg_Ticker'}
if not required_cols.issubset(df_ind.columns):
    raise KeyError(f"'{', '.join(required_cols)}' 칼럼이 모두 필요합니다. 파일을 확인해주세요.")

# 4) 티커별 조회 및 저장
# ───────────────────────────────────────────────────────────────────────────────
total = len(df_ind)
for idx, row in df_ind.iterrows():
    indicator_id = row['Indicator_ID'].strip()      # 지표 고유 ID
    ticker       = row['Bloomberg_Ticker'].strip()   # 블룸버그 티커

    print(f"[{idx+1}/{total}] 조회 시작 ▶ {indicator_id} ({ticker})")

    try:
        # PX_LAST(종가) 일별 시계열 조회
        df_ts = blp.bdh(
            tickers=[ticker],        # 리스트 형태로 입력
            flds='PX_LAST',          # 조회 필드: 종가
            start_date=start_date,   # 시작일
            end_date=end_date,       # 종료일
            Per='D',                 # 일별 데이터
            adjust='all'             # 배당·분할 등 조정 반영
        )
    except Exception as e:
        print(f"⚠️ 조회 실패: {indicator_id} ({ticker}) → {e}")
        continue

    if df_ts.empty:
        print(f"⚠️ 데이터 없음: {indicator_id} ({ticker}) 는 스킵합니다.")
        continue

    # 엑셀로 저장
    output_path = os.path.join(output_dir, f"{indicator_id}.xlsx")
    df_ts.to_excel(output_path, engine='openpyxl')
    print(f"✅ 저장 완료: {output_path}")

print("🎉 모든 지표 조회 및 저장이 완료되었습니다! 수고하셨습니다.")


# -*- coding: utf-8 -*-
"""
IND001 ~ IND151 엑셀 파일에 대해:
1) 시계열(날짜, 값) 자동 탐지
2) 최근 평균/표준편차 기반 임계치 계산
3) 마지막 값이 임계치를 기준으로
      - 정상       → C1 = 'G'
      - 주의/경고/심각 → C1 = 'Y'  (주의 이상은 모두 Y)
4) D1에 60일 평균(mu), E1에 60일 표준편차(sd=σ) 기록
5) 실행 요약 CSV 저장

설치:
    pip install pandas openpyxl numpy

사용:
    python script.py
"""

import os
import re
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Optional, Tuple

# ================= 팀장님 환경 설정 =================
BASE_DIR = r"C:\Users\amongpapa\chartup\go_scen\data\set"  # 엑셀 폴더 경로 (raw string 사용)
FILE_PREFIX = "IND"                                       # 파일 접두사
FILE_RANGE = range(1, 152)                                # 001~151

SHEET_NAME = 0                  # 첫 번째 시트(정수 또는 시트명)
WINDOW = 60                     # 롤링 윈도우(일수, 평균/표준편차 계산용)
K = 2.0                         # (현재는 z-score 참고용, 임계치에는 사용하지 않음)
MIN_PERIODS = max(30, WINDOW // 2)  # 최소 계산 데이터 길이(짧은 데이터 보호)

# 컬럼 이름 힌트(우선 매칭)
DATE_COL_HINTS = ["date", "날짜", "일자", "time", "일시"]
VALUE_COL_HINTS = ["close", "price", "value", "index", "지수", "종가", "가격", "값", "수치", "PX_LAST"]
# ====================================================

# ✅ [추가] 지표별 고정 임계치 설정 (방식 A: 절대값 기준)
# - 여기 있는 ID는 고정 임계치를 사용
# - 여기에 없는 ID는 자동 계산 임계치(평균 × 1.1 / 1.2 / 1.3)를 사용
CUSTOM_THRESHOLDS = {
    # USD/KRW 환율 (IND071): 값이 커질수록 리스크↑
    "IND071": {
        "direction": "up",   # up: 값이 커질수록 위험, down: 값이 작아질수록 위험
        "yellow": 1400.0,    # 주의
        "orange": 1600.0,    # 경고
        "red": 1800.0        # 심각
    },
    # 필요하면 코스피/미국 금리 등 추가 가능
    # "IND001": {
    #     "direction": "down",
    #     "yellow": 2400.0,
    #     "orange": 2200.0,
    #     "red": 2000.0
    # },
    # "IND050": {
    #     "direction": "up",
    #     "yellow": 4.5,
    #     "orange": 5.0,
    #     "red": 5.5
    # },
}


# ---------------- 도우미: 중복 컬럼명 유일화 ----------------
def _make_unique(names):
    """
    동일한 컬럼명이 반복될 경우 .1, .2 접미사를 부여해 유일화
    예) ['nan','nan'] -> ['nan','nan.1']
    """
    seen = {}
    out = []
    for n in names:
        key = str(n)
        if key not in seen:
            seen[key] = 0
            out.append(key)
        else:
            seen[key] += 1
            out.append(f"{key}.{seen[key]}")
    return out


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    공백 정리 + 유일화
    """
    cols = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    df.columns = _make_unique(cols)
    return df


# ---------------- 유틸: 숫자 문자열 → float (항상 Series 반환) ----------------
def _as_numeric_series(s) -> pd.Series:
    """
    어떤 입력(s: Series/DataFrame/ndarray/리스트)이 와도 **항상 pandas.Series**로 변환 후
    안전하게 숫자로 변환합니다.
    - 공백/대시류("", "-", "—", "_") → NaN
    - 괄호 음수 "(123)" → -123
    - 콤마 제거 "1,234.56" → 1234.56
    - 퍼센트 "5.2%" → 0.052
    """
    if isinstance(s, pd.DataFrame):
        # 숫자 변환률이 가장 높은 컬럼 하나 자동 선택
        best_col = None
        best_ratio = -1.0
        for col in s.columns:
            x = pd.to_numeric(
                pd.Series(s[col]).astype("string")
                .str.replace(",", "", regex=False)
                .str.replace("%", "", regex=False),
                errors="coerce",
            )
            ratio = x.notna().mean()
            if ratio > best_ratio:
                best_ratio, best_col = ratio, col
        s = s[best_col]

    if not isinstance(s, pd.Series):
        s = pd.Series(s)

    ss = s.astype("string").str.strip()
    ss = ss.replace(
        {
            "": pd.NA,
            "-": pd.NA,
            "—": pd.NA,
            "_": pd.NA,
            "nan": pd.NA,
            "NaN": pd.NA,
            "None": pd.NA,
        }
    )

    # 괄호 음수 처리
    neg_mask = ss.str.match(r"^\(.*\)$", na=False)
    ss2 = ss.str.replace(r"^\((.*)\)$", r"\1", regex=True)  # '(123)' -> '123'

    # 콤마/퍼센트 제거
    ss2 = ss2.str.replace(",", "", regex=False)
    pct_mask = ss2.str.endswith("%", na=False)
    ss2 = ss2.str.replace("%", "", regex=False)

    # 숫자 변환
    num = pd.to_numeric(ss2, errors="coerce")

    # 괄호 음수 적용
    if neg_mask.any():
        num.loc[neg_mask & num.notna()] = -num.loc[neg_mask & num.notna()].abs()

    # 퍼센트 → /100
    if pct_mask.any():
        num.loc[pct_mask & num.notna()] = num.loc[pct_mask & num.notna()] / 100.0

    return num.astype("float64")


# ---------------- 스코어링 헬퍼 ----------------
def _numeric_score(df: pd.DataFrame, lookahead: int = 5) -> int:
    """
    상위 lookahead행에서 숫자로 변환 가능한 열 개수 스코어링.
    """
    head = df.head(lookahead)
    cnt = 0
    for c in head.columns:
        s = head[c]
        s_num = _as_numeric_series(s)
        numlike = s_num.notna().mean()
        if numlike >= 0.5:
            cnt += 1
    return cnt


def _likely_good(df: pd.DataFrame) -> bool:
    """현재 형태가 '헤더-데이터'로 적절한지 간단 스코어."""
    return _numeric_score(df, lookahead=5) >= 2  # 숫자형 열 2개 이상이면 OK


# ---------------- 헤더 자동 감지 ----------------
def detect_header_and_read(path, sheet_name=0, max_scan=10) -> pd.DataFrame:
    """
    헤더/데이터 시작행이 불명확한 엑셀을 위한 로더.
    1) 일반 로드 후 간이 점검, 괜찮으면 그대로 사용 (컬럼 유일화 적용)
    2) 상단 max_scan행을 후보 헤더로 가정해 스코어링 후 베스트 헤더 채택
    """
    # 1) 일반 로드 시도
    df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=0)
    if df0.shape[0] == 0:
        df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
    if _likely_good(df0):
        return _clean_columns(df0)

    # 2) 후보 헤더 스캔
    raw = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
    nrows = min(max_scan, len(raw))
    best_score, best_row = -1e9, 0
    for hdr in range(nrows):
        tmp = raw.copy()
        tmp.columns = _make_unique(tmp.iloc[hdr].astype(str).str.strip())
        tmp = tmp.iloc[hdr + 1:].reset_index(drop=True)

        score_num = _numeric_score(tmp, lookahead=5)
        penalty = sum(
            1
            for name in tmp.columns
            if str(name).lower() in ("nan", "nat", "", "none")
        )
        score = score_num - 0.5 * penalty  # 나쁜 헤더 패널티

        if score > best_score:
            best_score, best_row = score, hdr

    raw.columns = _make_unique(raw.iloc[best_row].astype(str).str.strip())
    df = raw.iloc[best_row + 1:].reset_index(drop=True)
    return _clean_columns(df)


# ---------------- 날짜/값 컬럼 자동 선택 ----------------
def _pick_date_col(df: pd.DataFrame) -> Optional[str]:
    # 힌트 우선
    for hint in DATE_COL_HINTS:
        for c in df.columns:
            if hint.lower() in str(c).lower():
                return c
    # dtype/변환률 기반
    dt_cols = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
    if dt_cols:
        return dt_cols[0]
    best_col, best_ratio = None, 0.0
    for c in df.columns:
        try:
            converted = pd.to_datetime(df[c], errors="coerce", infer_datetime_format=True)
            ratio = converted.notna().mean()
            if ratio > best_ratio and ratio >= 0.7:
                best_col, best_ratio = c, ratio
        except Exception:
            pass
    return best_col


def _pick_value_col(df: pd.DataFrame, exclude_cols: list) -> Optional[str]:
    """
    값 컬럼 자동 선택(힌트 → 통계 스코어).
    """
    # 힌트 우선
    for hint in VALUE_COL_HINTS:
        for c in df.columns:
            if c in exclude_cols:
                continue
            if hint.lower() in str(c).lower():
                return c

    # 숫자형 후보 스코어링(결측↓, 분산>0, 샘플 수 충분)
    candidates = []
    for c in df.columns:
        if c in exclude_cols:
            continue
        series = _as_numeric_series(df[c])
        na_ratio = pd.isna(series).mean()
        var = np.nanvar(series.astype(float))
        if (~pd.isna(series)).sum() >= MIN_PERIODS and var > 0:
            candidates.append((c, na_ratio, var))

    if not candidates:
        # MIN_PERIODS 미만이어도 최선의 숫자형 컬럼을 fallback
        fallback = []
        for c in df.columns:
            if c in exclude_cols:
                continue
            series = _as_numeric_series(df[c])
            var = np.nanvar(series.astype(float))
            if var > 0:
                na_ratio = pd.isna(series).mean()
                fallback.append((c, na_ratio, var))
        if not fallback:
            return None
        fallback.sort(key=lambda x: (x[1], -x[2]))
        return fallback[0][0]

    candidates.sort(key=lambda x: (x[1], -x[2]))
    return candidates[0][0]


# ---------------- 로딩/계산 ----------------
def _load_timeseries(path: str, sheet_name=0) -> Tuple[pd.Series, pd.DataFrame]:
    """
    엑셀에서 시계열(날짜, 값)을 추출해 Series 반환.
    - 날짜 컬럼이 있으면 DatetimeIndex 정렬
    - 없으면 단순 순번 인덱스
    """
    df = detect_header_and_read(path, sheet_name=sheet_name)

    date_col = _pick_date_col(df)
    value_col = _pick_value_col(df, exclude_cols=[date_col] if date_col else [])

    if value_col is None:
        raise ValueError("숫자형 시계열 컬럼을 찾지 못했습니다. (헤더/형식 확인 필요)")

    if date_col is not None:
        dt = pd.to_datetime(df[date_col], errors="coerce", infer_datetime_format=True)
        df = df.loc[dt.notna()].copy()
        df.index = pd.to_datetime(df[date_col], errors="coerce")
        df = df.sort_index()
    else:
        df = df.reset_index(drop=True)

    s = _as_numeric_series(df[value_col]).astype(float).dropna()

    # 같은 날짜 중복 → 평균
    if isinstance(df.index, pd.DatetimeIndex) and s.index.has_duplicates:
        s = s.groupby(level=0).mean()

    return s.sort_index(), df


# ---------------- 임계치 기반 플래그 계산 ----------------
def compute_threshold_flag(
    s: pd.Series,
    indicator_id: Optional[str] = None,
    window: int = 60,
    k: float = 2.0,
) -> Tuple[str, dict]:
    """
    1) 60일 롤링 평균(mu), 표준편차(sd) 계산
    2) 임계치 결정
       - CUSTOM_THRESHOLDS에 등록된 ID: 고정 임계치 사용
       - 그 외: mu×1.1, 1.2, 1.3 자동 임계치 사용
    3) 마지막 값의 레벨 판단
       - level ∈ {normal, yellow, orange, red}
       - 팀장님 요청: 주의 이상(yellow, orange, red)은 모두 'Y', 그 외는 'G'
    반환:
      flag: 'G' 또는 'Y'
      info: 각종 참고 정보(dict)
    """
    if len(s) < MIN_PERIODS:
        # 데이터가 너무 짧으면 보수적으로 G 처리
        return "G", {"reason": "too_short", "len": len(s)}

    roll_mean = s.rolling(window=window, min_periods=MIN_PERIODS).mean()
    roll_std = s.rolling(window=window, min_periods=MIN_PERIODS).std(ddof=0)

    last_val = s.iloc[-1]
    mu = roll_mean.iloc[-1]
    sd = roll_std.iloc[-1]

    if pd.isna(mu) or pd.isna(sd) or sd == 0:
        return "G", {
            "reason": "nan_or_zero_std",
            "mu": mu,
            "sd": sd,
            "last": last_val,
        }

    # z-score는 참고용으로만 계산
    z = (last_val - mu) / sd

    # 1) 임계치 결정 (고정/자동)
    ind_id = (indicator_id or "").upper()
    cfg = CUSTOM_THRESHOLDS.get(ind_id)

    if cfg is not None:
        # ✅ 고정 임계치 사용
        direction = cfg.get("direction", "up")
        thr_yellow = cfg["yellow"]
        thr_orange = cfg["orange"]
        thr_red = cfg["red"]
        reason = "custom_threshold"
    else:
        # ✅ 자동 임계치 (평균 × 비율)
        direction = "up"  # 기본: 값이 올라갈수록 위험한 지표로 가정
        thr_yellow = mu * 1.1
        thr_orange = mu * 1.2
        thr_red = mu * 1.3
        reason = "auto_threshold"

    # 2) 레벨 판정
    if direction == "up":
        if last_val >= thr_red:
            level = "red"
        elif last_val >= thr_orange:
            level = "orange"
        elif last_val >= thr_yellow:
            level = "yellow"
        else:
            level = "normal"
    else:  # direction == "down": 값이 작아질수록 위험
        if last_val <= thr_red:
            level = "red"
        elif last_val <= thr_orange:
            level = "orange"
        elif last_val <= thr_yellow:
            level = "yellow"
        else:
            level = "normal"

    # 3) 팀장님 요청: "주의 이상은 Y" → normal만 G, 나머지는 전부 Y
    if level == "normal":
        flag = "G"
    else:
        flag = "Y"

    info = {
        "last": last_val,
        "mu": mu,
        "sd": sd,
        "zscore": z,
        "thr_yellow": thr_yellow,
        "thr_orange": thr_orange,
        "thr_red": thr_red,
        "level": level,
        "reason": reason,
    }

    # 예전 구조 호환용 필드(upper/lower)는 None으로 둠
    info["upper"] = None
    info["lower"] = None

    return flag, info


# ---------------- 엑셀 기록 ----------------
def write_results_to_excel(
    path: str,
    flag: str,
    mu: Optional[float],
    sd: Optional[float],
    sheet_name=0,
):
    """
    엑셀 파일의 C1/D1/E1 업데이트:
      - C1: G 또는 Y (주의 이상은 모두 Y)
      - D1: 60일 롤링 평균(mu)
      - E1: 60일 롤링 표준편차(sd = σ)
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

    ws["C1"] = flag

    def _num_or_none(x):
        try:
            if x is None:
                return None
            if isinstance(x, (int, float)) and not (
                isinstance(x, float) and np.isnan(x)
            ):
                return float(x)
            return None
        except Exception:
            return None

    ws["D1"] = _num_or_none(mu)  # 60일 평균
    ws["E1"] = _num_or_none(sd)  # 60일 표준편차(σ)

    wb.save(path)


# ---------------- 메인 드라이버 ----------------
def main():
    results = []
    for i in FILE_RANGE:
        fname = f"{FILE_PREFIX}{i:03d}.xlsx"
        fpath = os.path.join(BASE_DIR, fname)
        if not os.path.exists(fpath):
            results.append((fname, "missing", None))
            print(f"[SKIP] {fname} (file not found)")
            continue

        try:
            s, _df = _load_timeseries(fpath, sheet_name=SHEET_NAME)

            # ✅ [변경] 지표 ID를 넘겨서 임계치 기반 플래그 계산
            indicator_id = f"{FILE_PREFIX}{i:03d}"
            flag, info = compute_threshold_flag(
                s, indicator_id=indicator_id, window=WINDOW, k=K
            )

            mu = info.get("mu") if isinstance(info, dict) else None
            sd = info.get("sd") if isinstance(info, dict) else None
            write_results_to_excel(fpath, flag, mu, sd, sheet_name=SHEET_NAME)

            ztxt = (
                f"{info.get('zscore'):.2f}"
                if isinstance(info, dict)
                and "zscore" in info
                and pd.notna(info.get("zscore"))
                else "n/a"
            )
            mutxt = (
                f"{mu:.6g}"
                if isinstance(mu, (int, float)) and not pd.isna(mu)
                else "n/a"
            )
            sdtxt = (
                f"{sd:.6g}"
                if isinstance(sd, (int, float)) and not pd.isna(sd)
                else "n/a"
            )

            level = info.get("level") if isinstance(info, dict) else "n/a"

            print(
                f"[OK] {fname} -> C1='{flag}', level={level}, "
                f"D1(mu)={mutxt}, E1(sd)={sdtxt} (z={ztxt})"
            )
            results.append((fname, flag, info))
        except Exception as e:
            err_msg = str(e)
            print(f"[ERR] {fname}: {err_msg}")
            # 간단 프로브: 컬럼/타입 힌트 출력 (헤더/형식 추적)
            try:
                df_probe = detect_header_and_read(fpath, sheet_name=SHEET_NAME)
                print(f"  -> columns: {list(df_probe.columns)}")
                sample_info = []
                for c in df_probe.columns[:10]:
                    sample_info.append(f"{c}:{str(df_probe[c].dtype)}")
                print("  -> dtypes(head):", ", ".join(sample_info))
            except Exception as e2:
                print(f"  -> probe failed: {e2}")
            results.append((fname, "error", err_msg))

    # 요약 CSV 저장
    summary_path = os.path.join(
        BASE_DIR,
        f"vol_band_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
    )
    rows = []
    for fname, flag, info in results:
        row = {"file": fname, "flag": flag}
        if isinstance(info, dict):
            row.update(
                {
                    "last": info.get("last"),
                    "mu": info.get("mu"),
                    "sd": info.get("sd"),
                    "upper": info.get("upper"),
                    "lower": info.get("lower"),
                    "zscore": info.get("zscore"),
                    "thr_yellow": info.get("thr_yellow"),
                    "thr_orange": info.get("thr_orange"),
                    "thr_red": info.get("thr_red"),
                    "level": info.get("level"),
                    "reason": info.get("reason"),
                }
            )
        elif isinstance(info, str):
            row["error"] = info
        rows.append(row)
    pd.DataFrame(rows).to_csv(summary_path, index=False, encoding="utf-8-sig")
    print(f"\n요약 저장: {summary_path}")


if __name__ == "__main__":
    main()


# -*- coding: utf-8 -*-
r"""
IND001~IND151 엑셀만 사용하여 파생지표를 계산하고
id, data(=JSON 문자열) 2-컬럼 형식으로 하나의 엑셀 파일에 저장합니다.

저장 위치:
  C:\Users\amongpapa\chartup\go_scen\data\market_data\market_data.xlsx

설치:
  pip install pandas numpy openpyxl
"""

import os
import re
import json
import math
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Optional, Tuple, Dict

# ================= 팀장님 환경 설정 =================
BASE_DIR   = r"C:\Users\amongpapa\chartup\go_scen\data\set"                  # IND 엑셀 폴더
OUT_PATH   = r"C:\Users\amongpapa\chartup\go_scen\data\market_data\market_data.xlsx"
FILE_PREFIX = "IND"
FILE_RANGE  = range(1, 152)                                                  # 001~151
SHEET_NAME  = 0

WINDOWS_RET = [20, 60, 120]
ANNUAL_DAYS = 252
VOL_QUANTILES = (0.33, 0.66)

# 상관/베타용 벤치마크 매핑(필요 시 채우세요: 예 'KOSPI': 'IND001')
BENCHMARK_MAP: Dict[str, str] = {}

DATE_COL_HINTS  = ["date", "날짜", "일자", "time", "일시"]
VALUE_COL_HINTS = ["close", "price", "value", "index", "지수", "종가", "가격", "값", "수치"]

MIN_PERIODS = 30  # 짧은 데이터 보호
# ====================================================


# ---------------- 유틸: 엑셀 헤더 자동 감지 ----------------
def detect_header_and_read(path, sheet_name=0, max_scan=10) -> pd.DataFrame:
    """
    헤더/데이터 시작행 불명확 엑셀을 안전하게 읽기:
      1) 일반 로드 후 품질 점검, 통과시 그대로 사용
      2) 위에서부터 max_scan줄을 헤더 후보로 스코어링해 최적 헤더 채택
    """
    try:
        df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=0)
    except Exception:
        df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)

    if df0.shape[0] == 0:
        df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)

    if _likely_good(df0):
        return _clean_columns(df0)

    raw = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
    nrows = min(max_scan, len(raw))
    best_score, best_row = -1, 0
    for hdr in range(nrows):
        tmp = raw.copy()
        tmp.columns = tmp.iloc[hdr].astype(str).str.strip()
        tmp = tmp.iloc[hdr + 1 :].reset_index(drop=True)
        score = _numeric_score(tmp, lookahead=5)
        if score > best_score:
            best_score, best_row = score, hdr

    raw.columns = raw.iloc[best_row].astype(str).str.strip()
    df = raw.iloc[best_row + 1 :].reset_index(drop=True)
    return _clean_columns(df)

def _likely_good(df: pd.DataFrame) -> bool:
    return _numeric_score(df, lookahead=5) >= 2

def _numeric_score(df: pd.DataFrame, lookahead: int = 5) -> int:
    head = df.head(lookahead)
    cnt = 0
    for c in head.columns:
        s = head[c]
        s_num = _as_numeric_series(s, strict=False)
        if s_num.notna().mean() >= 0.5:
            cnt += 1
    return cnt

def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    return df


# ---------------- 숫자 문자열 → float 시리즈(강력 방어) ----------------
def _as_numeric_series(s: pd.Series, strict: bool = True) -> pd.Series:
    """
    문자 섞인 숫자를 float로 정제(%, 괄호음수, 콤마 등)하여 항상 Series 반환.
    strict=False 이면 예외는 조용히 NaN으로 보냄.
    """
    try:
        # 이미 숫자형이면 조용히 변환
        if pd.api.types.is_numeric_dtype(s):
            return pd.to_numeric(s, errors="coerce")

        # 어떤 타입이든 문자열 dtype으로 강제(Series.str 사용 보장)
        ss = s.astype("string")  # pandas StringDtype
        ss = ss.str.strip()

        # 빈값/대시류 → NaN
        ss = ss.replace({"": pd.NA, "-": pd.NA, "—": pd.NA, "_": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})

        # 괄호 음수
        neg_mask = ss.str.match(r"^\(.*\)$", na=False)
        ss2 = ss.str.replace(r"^\((.*)\)$", r"\1", regex=True)

        # 콤마 제거, % 제거
        ss2 = ss2.str.replace(",", "", regex=False)
        pct_mask = ss2.str.endswith("%", na=False)
        ss2 = ss2.str.replace("%", "", regex=False)

        num = pd.to_numeric(ss2, errors="coerce")

        # 괄호 음수 적용
        if neg_mask.any():
            idx = neg_mask & num.notna()
            num.loc[idx] = -num.loc[idx].abs()

        # 퍼센트 → /100
        if pct_mask.any():
            idx = pct_mask & num.notna()
            num.loc[idx] = num.loc[idx] / 100.0

        # 항상 Series
        if not isinstance(num, pd.Series):
            num = pd.Series(num, index=s.index, dtype="float64")
        return num.astype(float)

    except Exception:
        if strict:
            raise
        # 문제 있으면 전부 NaN으로 리턴(스코어링용)
        return pd.to_numeric(pd.Series([pd.NA]*len(s), index=s.index), errors="coerce")


# ---------------- 날짜/값 컬럼 자동 선택 ----------------
def _pick_date_col(df: pd.DataFrame) -> Optional[str]:
    for hint in DATE_COL_HINTS:
        for c in df.columns:
            if hint.lower() in str(c).lower():
                return c
    dt_cols = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
    if dt_cols:
        return dt_cols[0]
    best_col, best_ratio = None, 0.0
    for c in df.columns:
        try:
            converted = pd.to_datetime(df[c], errors="coerce", infer_datetime_format=True)
            ratio = converted.notna().mean()
            if ratio > best_ratio and ratio >= 0.7:
                best_col, best_ratio = c, ratio
        except Exception:
            pass
    return best_col

def _pick_value_col(df: pd.DataFrame, exclude_cols: list) -> Optional[str]:
    # 힌트 우선
    for hint in VALUE_COL_HINTS:
        for c in df.columns:
            if c in exclude_cols: 
                continue
            if hint.lower() in str(c).lower():
                return c

    # 숫자형 후보 스코어링
    candidates = []
    for c in df.columns:
        if c in exclude_cols:
            continue
        try:
            series = _as_numeric_series(df[c], strict=False)
            na_ratio = pd.isna(series).mean()
            var = np.nanvar(series.astype(float))
            if (~pd.isna(series)).sum() >= max(MIN_PERIODS, 10) and var > 0:
                candidates.append((c, na_ratio, var))
        except Exception:
            continue

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[1], -x[2]))  # 결측↓, 분산↑
    return candidates[0][0]


# ---------------- 시계열 로딩 ----------------
def load_series_from_excel(path: str, sheet_name=0) -> pd.Series:
    """엑셀에서 (날짜, 값) 시계열 Series 반환 (DatetimeIndex 정렬)"""
    df = detect_header_and_read(path, sheet_name=sheet_name)

    date_col = _pick_date_col(df)
    value_col = _pick_value_col(df, exclude_cols=[date_col] if date_col else [])
    if value_col is None:
        raise ValueError("숫자형 시계열 컬럼을 찾지 못했습니다.")

    if date_col is not None:
        dt = pd.to_datetime(df[date_col], errors="coerce", infer_datetime_format=True)
        df = df.loc[dt.notna()].copy()
        df.index = pd.to_datetime(df[date_col], errors="coerce")
        df = df.sort_index()
    else:
        df = df.reset_index(drop=True)

    s = _as_numeric_series(df[value_col]).astype(float).dropna()

    # 같은 날짜 중복 → 평균
    if isinstance(df.index, pd.DatetimeIndex) and s.index.has_duplicates:
        s = s.groupby(level=0).mean()

    s = s.sort_index()
    if len(s) < MIN_PERIODS:
        raise ValueError(f"데이터 길이 부족(len={len(s)})")

    return s


# ---------------- 보조 계산 함수 ----------------
def pct_return(s: pd.Series, days: int) -> Optional[float]:
    if len(s) <= days:
        return None
    return float(s.iloc[-1] / s.iloc[-days-1] - 1.0)

def ann_return_from_period(r: Optional[float], days: int) -> Optional[float]:
    if r is None:
        return None
    try:
        return float((1.0 + r) ** (ANNUAL_DAYS / days) - 1.0)
    except Exception:
        return None

def rolling_vol_annualized(s: pd.Series, days: int) -> Optional[float]:
    if len(s) < max(days+1, MIN_PERIODS):
        return None
    rets = s.pct_change()
    vol = rets.rolling(window=days, min_periods=int(days*0.6)).std(ddof=0)
    last_vol = vol.iloc[-1]
    if pd.isna(last_vol):
        return None
    return float(last_vol * np.sqrt(ANNUAL_DAYS))

def drawdown_stats(s: pd.Series, lookback_days: int) -> Tuple[Optional[float], Optional[float]]:
    if len(s) < max(lookback_days, MIN_PERIODS):
        return (None, None)
    sub = s.iloc[-lookback_days:]
    peak = sub.cummax()
    dd = sub / peak - 1.0
    mdd = dd.min()
    curr_dd = dd.iloc[-1]
    return (float(mdd), float(curr_dd))

def sma(s: pd.Series, days: int) -> Optional[float]:
    if len(s) < days:
        return None
    return float(s.rolling(days).mean().iloc[-1])

def ema(s: pd.Series, days: int) -> Optional[float]:
    if len(s) < days:
        return None
    return float(s.ewm(span=days, adjust=False).mean().iloc[-1])

def gap_pct(price: float, ref: Optional[float]) -> Optional[float]:
    if ref is None or ref == 0 or price is None:
        return None
    return float(price / ref - 1.0)

def golden_dead_cross_state(s: pd.Series, short=20, long=60) -> Optional[str]:
    if len(s) < long + 1:
        return None
    sma_s = s.rolling(short).mean()
    sma_l = s.rolling(long).mean()
    prev = np.sign(sma_s.iloc[-2] - sma_l.iloc[-2])
    curr = np.sign(sma_s.iloc[-1] - sma_l.iloc[-1])
    if prev <= 0 and curr > 0:
        return "golden"
    elif prev >= 0 and curr < 0:
        return "dead"
    else:
        return "none"

def slope_tstat_lastN(s: pd.Series, N: int = 60) -> Tuple[Optional[float], Optional[float]]:
    if len(s) < N:
        return (None, None)
    y = s.iloc[-N:].values.astype(float)
    x = np.arange(N).astype(float)
    x_mean = x.mean()
    y_mean = y.mean()
    cov_xy = np.sum((x - x_mean) * (y - y_mean))
    var_x = np.sum((x - x_mean) ** 2)
    if var_x == 0:
        return (None, None)
    beta1 = cov_xy / var_x
    y_hat = (beta1 * (x - x_mean)) + y_mean
    resid = y - y_hat
    s2 = np.sum(resid**2) / (N - 2)
    se_beta1 = math.sqrt(s2 / var_x)
    tstat = beta1 / se_beta1 if se_beta1 != 0 else None
    return (float(beta1), float(tstat) if tstat is not None else None)

def pos_in_52w(s: pd.Series) -> Optional[float]:
    if len(s) < 20:
        return None
    look = s.iloc[-min(len(s), 252):]
    hi, lo = float(look.max()), float(look.min())
    last = float(look.iloc[-1])
    if hi == lo:
        return None
    return float((last - lo) / (hi - lo))

def vol_regime_label(s: pd.Series, days: int = 60) -> Optional[str]:
    if len(s) < max(days+20, MIN_PERIODS+20):
        return None
    rets = s.pct_change()
    vol = rets.rolling(window=days, min_periods=int(days*0.6)).std(ddof=0) * np.sqrt(ANNUAL_DAYS)
    curr = vol.iloc[-1]
    hist = vol.dropna().values
    if np.isnan(curr) or len(hist) < 30:
        return None
    q1, q2 = np.quantile(hist, VOL_QUANTILES)
    if curr < q1:
        return "low"
    elif curr < q2:
        return "mid"
    else:
        return "high"

def trend_label_from_slope(s: pd.Series, N: int = 60) -> Optional[str]:
    slope, t = slope_tstat_lastN(s, N)
    if slope is None or t is None:
        return None
    if t > 2:
        return "up"
    elif t < -2:
        return "down"
    else:
        return "flat"

def compute_beta(x: pd.Series, mkt: pd.Series, days: int = 60) -> Optional[float]:
    if len(x) < days+1 or len(mkt) < days+1:
        return None
    rx = x.pct_change().iloc[-days:]
    rm = mkt.pct_change().iloc[-days:]
    df = pd.concat([rx, rm], axis=1).dropna()
    if len(df) < int(days*0.6):
        return None
    cov = np.cov(df.iloc[:,0], df.iloc[:,1])[0,1]
    var = np.var(df.iloc[:,1])
    if var == 0:
        return None
    return float(cov / var)

def corr_rolling(x: pd.Series, y: pd.Series, days: int = 60) -> Optional[float]:
    if len(x) < days+1 or len(y) < days+1:
        return None
    rx = x.pct_change()
    ry = y.pct_change()
    corr = rx.rolling(days).corr(ry)
    val = corr.iloc[-1]
    return float(val) if pd.notna(val) else None


# ---------------- 메인 파이프라인 ----------------
def main():
    series_map: Dict[str, pd.Series] = {}
    for i in FILE_RANGE:
        fid = f"{FILE_PREFIX}{i:03d}"
        fpath = os.path.join(BASE_DIR, f"{fid}.xlsx")
        if not os.path.exists(fpath):
            print(f"[SKIP] {fid}: file not found")
            continue
        try:
            s = load_series_from_excel(fpath, sheet_name=SHEET_NAME)
            series_map[fid] = s
        except Exception as e:
            # 핵심: 문자열 처리/컬럼매핑 실패 시에도 원인 출력하고 계속
            print(f"[SKIP] {fid}: {e}")

    if not series_map:
        raise RuntimeError("유효한 IND 시계열을 하나도 찾지 못했습니다.")

    # 벤치마크 해석(옵션)
    bench_series: Dict[str, pd.Series] = {}
    for name, ind in BENCHMARK_MAP.items():
        if ind in series_map:
            bench_series[name] = series_map[ind]

    rows = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for fid, s in series_map.items():
        last = float(s.iloc[-1])
        asof = str(s.index[-1]) if isinstance(s.index, pd.DatetimeIndex) else now_str

        # 변동성/수익률
        rets = {}; rets_ann = {}; vols = {}
        for d in WINDOWS_RET:
            r = pct_return(s, d)
            rets[f"{d}d"] = r
            rets_ann[f"{d}d_ann"] = ann_return_from_period(r, d)
            vols[f"{d}d_vol_ann"] = rolling_vol_annualized(s, d)

        # 성과(YTD/1M/3M/6M/1Y)
        period_map = {"1m":21, "3m":63, "6m":126, "1y":252}
        perf = {k: pct_return(s, d) for k, d in period_map.items()}
        if isinstance(s.index, pd.DatetimeIndex):
            year = s.index[-1].year
            ytd_slice = s[s.index.year == year]
            ytd_first = ytd_slice.iloc[0] if len(ytd_slice) > 0 else None
            perf["ytd"] = float(last / ytd_first - 1.0) if ytd_first and ytd_first != 0 else None
        else:
            perf["ytd"] = None

        # 낙폭
        mdd6, dd6   = drawdown_stats(s, 126)
        mdd1y, dd1y = drawdown_stats(s, 252)

        # 추세/괴리
        sma20, sma60, sma120 = sma(s,20), sma(s,60), sma(s,120)
        ema20, ema60, ema120 = ema(s,20), ema(s,60), ema(s,120)
        gaps = {
            "gap_to_sma20":  gap_pct(last, sma20),
            "gap_to_sma60":  gap_pct(last, sma60),
            "gap_to_sma120": gap_pct(last, sma120),
            "gap_to_ema20":  gap_pct(last, ema20),
            "gap_to_ema60":  gap_pct(last, ema60),
            "gap_to_ema120": gap_pct(last, ema120),
        }
        cross_20_60  = golden_dead_cross_state(s, 20, 60)
        cross_60_120 = golden_dead_cross_state(s, 60, 120)

        # 기울기/t
        slope60, t60 = slope_tstat_lastN(s, 60)

        # 52주 위치
        pos52w = pos_in_52w(s)
        pos52w_pct = float(pos52w*100.0) if pos52w is not None else None

        # 레짐
        vol_reg   = vol_regime_label(s, 60)
        trend_reg = trend_label_from_slope(s, 60)
        regime    = f"{trend_reg}-{vol_reg}" if trend_reg and vol_reg else None

        # 상관/베타(옵션)
        corr_res = {}
        beta_res = {}
        for bname, bs in bench_series.items():
            corr_res[bname] = corr_rolling(s, bs, 60)
        mkt_key = next((k for k in ["KOSPI", "S&P500", "KOSDAQ"] if k in bench_series), None)
        if mkt_key:
            beta_res[mkt_key] = compute_beta(s, bench_series[mkt_key], 60)

        # 요약 최근값
        latest = {"last": last, "asof": asof, "ytd": perf.get("ytd"),
                  "1m": perf.get("1m"), "3m": perf.get("3m"),
                  "6m": perf.get("6m"), "1y": perf.get("1y"), "note": None}

        # === id, data(JSON) 저장 ===
        def add_row(suffix: str, obj):
            rows.append({"id": f"{fid}:{suffix}", "data": json.dumps(obj, ensure_ascii=False)})

        add_row("returns", {**rets, **rets_ann})
        add_row("vol", vols)
        add_row("mdd", {"mdd_6m": mdd6, "dd_6m": dd6, "mdd_1y": mdd1y, "dd_1y": dd1y})
        add_row("trend",
                {"sma20":sma20,"sma60":sma60,"sma120":sma120,
                 "ema20":ema20,"ema60":ema60,"ema120":ema120,
                 **gaps,
                 "cross_20_60": cross_20_60, "cross_60_120": cross_60_120})
        add_row("slope60", {"slope": slope60, "tstat": t60})
        add_row("pos_52w", {"pos_0to1": pos52w, "pos_pct": pos52w_pct})
        add_row("regime", {"trend": trend_reg, "vol": vol_reg, "regime": regime})
        if corr_res:
            add_row("corr60", corr_res)
        if beta_res:
            add_row("beta60", beta_res)
        add_row("latest", latest)

        summary = {
            "latest": latest,
            "returns": {**rets, **rets_ann},
            "vol": vols,
            "mdd": {"mdd_6m": mdd6, "dd_6m": dd6, "mdd_1y": mdd1y, "dd_1y": dd1y},
            "trend": {"gaps": gaps, "cross_20_60": cross_20_60, "cross_60_120": cross_60_120},
            "slope60": {"slope": slope60, "tstat": t60},
            "pos_52w_pct": pos52w_pct,
            "regime": {"trend": trend_reg, "vol": vol_reg, "regime": regime},
            "corr60": corr_res if corr_res else None,
            "beta60": beta_res if beta_res else None
        }
        add_row("summary", summary)

    # 저장
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    df_out = pd.DataFrame(rows, columns=["id", "data"])
    df_out.to_excel(OUT_PATH, index=False, engine="openpyxl")
    print(f"[OK] 저장 완료: {OUT_PATH} (rows={len(df_out)})")


if __name__ == "__main__":
    main()


# -*- coding: utf-8 -*-
"""
목적:
- C:/Users/amongpapa/chartup/raw_data 폴더의 risk_thresholds_YYYYMMDD.xlsx 파일들 중,
  파일명 날짜(YYYYMMDD)가 '가장 최신'인 파일을 자동 선택
- 해당 파일의 alerts 시트를 읽어서,
  엑셀 2행 → IND500, 3행 → IND501, ... 규칙으로 매핑
- 각 행에 'TRUE'가 한 칸이라도 있으면 C1='Y', 아니면 'G'를
  C:/Users/amongpapa/chartup/go_scen/data/set/IND###.xlsx 에 기록
- 대상 범위: IND500 ~ IND700 (존재하는 파일만 업데이트)

설치:
    pip install pandas openpyxl

팀장님 환경 유의:
- 윈도우 경로는 pathlib.Path(r"...") 형식 사용
- 슬래시(/)를 써도 됩니다. 여기서는 Path를 사용하므로 안전합니다.
"""

from pathlib import Path  # ✅ 윈도우에서도 안전한 경로 처리
from datetime import datetime
import re
import pandas as pd
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────
# 1) 경로/파라미터 설정
# ─────────────────────────────────────────────────────────────────────
# risk_thresholds 엑셀들이 있는 폴더
RISK_DIR = Path(r"C:\Users\amongpapa\chartup\raw_data")

# alerts 시트명
ALERTS_SHEET = "alerts"

# IND 엑셀들이 있는 폴더(업데이트 대상)
TARGET_DIR = Path(r"C:\Users\amongpapa\chartup\go_scen\data\set")

# 매핑: 엑셀 '2행' → IND500 (즉, row 2 => 500, row 3 => 501, ...)
START_EXCEL_ROW = 2
START_IND = 500
MAX_IND = 700  # 이 값까지 존재하는 파일만 업데이트

# 업데이트할 워크시트: 첫 시트를 사용(정수 0). 특정 시트명이면 문자열로 지정 가능.
TARGET_SHEET = 0

# 'TRUE' 판정 시 포함할 값(대소문자, 공백 대응)
TRUE_TOKENS = {"TRUE"}  # 문자열 'TRUE' 또는 불리언 True를 허용
# ─────────────────────────────────────────────────────────────────────


# ---------------- 최신 파일 선택: 파일명 날짜(YYYYMMDD) 기준 ----------------
def pick_risk_file_by_name_date() -> Path:
    """
    risk_thresholds_YYYYMMDD.xlsx 파일들에서 '파일명 날짜'가 가장 최신인 파일을 선택.
    예: risk_thresholds_20251006.xlsx 가 있으면 그걸 사용.
    """
    patt = re.compile(r"^risk_thresholds_(\d{8})\.xlsx$", re.IGNORECASE)
    best = None
    best_dt = None
    for p in RISK_DIR.glob("risk_thresholds_*.xlsx"):
        m = patt.match(p.name)
        if not m:
            continue
        try:
            dt = datetime.strptime(m.group(1), "%Y%m%d")
        except ValueError:
            continue
        if (best_dt is None) or (dt > best_dt):
            best_dt = dt
            best = p
    if best is None:
        raise FileNotFoundError(f"폴더에 risk_thresholds_YYYYMMDD.xlsx 파일이 없습니다: {RISK_DIR}")
    return best


# ---------------- 엑셀 행 → IND 번호 매핑 ----------------
def excel_row_to_ind(excel_row: int) -> int:
    """
    엑셀 실제 행 번호(1-based) → IND 번호 매핑
    예) 2행 → 500, 3행 → 501 ...
    """
    return START_IND + (excel_row - START_EXCEL_ROW)


# ---------------- 한 행에 TRUE 존재 여부 검사 ----------------
def row_has_true(row_values) -> bool:
    """
    한 행에 TRUE가 하나라도 있으면 True 반환.
    - 불리언 True
    - 문자열 'TRUE' (대소문자 무시, 앞뒤 공백 무시)
    """
    for v in row_values:
        # 불리언 True
        if v is True:
            return True
        # 문자열 'TRUE'
        if isinstance(v, str) and v.strip().upper() in TRUE_TOKENS:
            return True
    return False


# ---------------- IND 엑셀의 C1 업데이트 ----------------
def update_c1_flag(ind_num: int, flag: str) -> bool:
    """
    IND 파일의 C1을 flag('Y'/'G')로 업데이트.
    - 파일이 없으면 False 반환(스킵)
    - 성공하면 True
    """
    fpath = TARGET_DIR / f"IND{ind_num:03d}.xlsx"
    if not fpath.exists():
        print(f"  ⏭️ 스킵 (파일없음): {fpath.name}")
        return False

    wb = load_workbook(fpath)
    ws = wb.worksheets[TARGET_SHEET] if isinstance(TARGET_SHEET, int) else wb[TARGET_SHEET]
    ws["C1"] = flag  # C1만 기록 (D1/E1은 변경하지 않음)
    wb.save(fpath)
    print(f"  ✅ C1='{flag}' 기록: {fpath.name}")
    return True


# ---------------- 메인 루틴 ----------------
def main():
    # 1) 최신 risk_thresholds 파일 선택(파일명 날짜 기준)
    risk_file = pick_risk_file_by_name_date()
    print(f"▶ 기준 파일(파일명 날짜 최신): {risk_file.name}")

    # 2) alerts 시트를 헤더 없이(raw) 로드 → 엑셀 실제 행 번호와 1:1 매핑 가능
    df = pd.read_excel(risk_file, sheet_name=ALERTS_SHEET, header=None, engine="openpyxl")
    nrows = df.shape[0]

    total_updated = 0
    total_skipped = 0

    # 3) 엑셀 '2행'부터 끝까지 순회하며 IND 번호로 매핑
    #    단, IND 번호가 MAX_IND를 넘으면 중단
    for excel_row in range(START_EXCEL_ROW, nrows + 1):
        ind_num = excel_row_to_ind(excel_row)

        if ind_num < START_IND:
            continue
        if ind_num > MAX_IND:
            break  # 범위를 넘으면 종료

        row_vals = df.iloc[excel_row - 1].tolist()  # pandas는 0-based → excel_row-1
        has_true = row_has_true(row_vals)
        flag = "Y" if has_true else "G"

        print(f"[행 {excel_row} → IND{ind_num:03d}] TRUE존재={has_true} → C1='{flag}'")
        ok = update_c1_flag(ind_num, flag)
        if ok:
            total_updated += 1
        else:
            total_skipped += 1

    print("\n──────── 요약 ────────")
    print(f"업데이트 성공: {total_updated}개")
    print(f"스킵(파일없음 등): {total_skipped}개")
    print("완료 ✅")


if __name__ == "__main__":
    main()


import pandas as pd
from pathlib import Path

# 📂 파일 경로
market_data_path = Path(r"C:\Users\amongpapa\chartup\go_scen\data\market_data\market_data.xlsx")
indicator_path   = Path(r"C:\Users\amongpapa\lm\agent\go_scen\data\indicator.xlsx")

# 1️⃣ 엑셀 불러오기
market_df = pd.read_excel(market_data_path)
indicator_df = pd.read_excel(indicator_path)

# 2️⃣ id 칼럼에서 "IND001:returns" → "IND001" 부분 추출
market_df["Indicator_ID"] = market_df["id"].str.split(":").str[0]

# 3️⃣ indicator.xlsx 매핑
merged_df = market_df.merge(
    indicator_df[["Indicator_ID", "Indicator_Name", "Bloomberg_Ticker"]],
    on="Indicator_ID",
    how="left"
)

# 4️⃣ C, D열 위치에 Indicator_Name, Bloomberg_Ticker 삽입
cols = list(market_df.columns)
insert_position = 2  # C열 위치
for new_col in ["Indicator_Name", "Bloomberg_Ticker"]:
    cols.insert(insert_position, new_col)
    insert_position += 1

final_df = merged_df[cols]

# 5️⃣ 기존 파일 덮어쓰기 (업데이트 저장)
final_df.to_excel(market_data_path, index=False)

print(f"[OK] market_data.xlsx 파일이 업데이트되었습니다: {market_data_path}")


# -*- coding: utf-8 -*-
"""
폴더 비우기 유틸 (보존 예외 지원, **폴더는 절대 삭제하지 않음**)
- target 폴더 내부의 파일/하위폴더 '내용'만 삭제하되, 지정한 예외 파일/폴더는 보존
- dry_run=True 시 실제 삭제 없이 로그만 출력
- Windows에서도 안전하게: pathlib.Path 사용 + / 슬래시 경로
"""

from pathlib import Path
import shutil  # (남겨둠: 추후 확장 대비, 현재 폴더 삭제에는 사용하지 않음)
from typing import Iterable, Set

# [설정] 대상 폴더들
TARGET_BIGKINDS = Path(r"C:/Users/amongpapa/chartup/go_scen/data/news/bigkinds")
TARGET_DAILY_NEWS = Path(r"C:/Users/amongpapa/chartup/go_scen/data/news/bigkinds/daily_news")

def _is_dangerous_path(p: Path) -> bool:
    """
    루트 같은 위험 경로를 보호하기 위한 간단 가드.
    - 드라이브 루트(ex: C:/) 또는 경로 문자열이 너무 짧은 경우 위험으로 간주
    """
    try:
        p = p.resolve()
    except Exception:
        return True
    return (p == Path(p.anchor)) or (len(str(p)) < 10)

def purge_dir(target: Path, dry_run: bool = False, exclude: Iterable[Path] = ()) -> None:
    """
    target 폴더 안의 모든 '내용물'을 정리(파일/링크 삭제)하되, **폴더 자체는 삭제하지 않음**.
    exclude는 절대/상대 경로 모두 허용:
      - 절대경로: 그대로 보존
      - 상대경로: target 기준 상대 경로로 해석하여 보존
    """
    target = Path(target).resolve()
    if not target.is_dir():
        raise FileNotFoundError(f"대상 폴더를 찾을 수 없습니다: {target}")

    if _is_dangerous_path(target):
        raise ValueError(f"위험한 경로로 판단되어 중단합니다: {target}")

    # 보존(제외) 집합: 절대 경로로 정규화
    exclude_abs: Set[Path] = set()
    for ex in exclude:
        ex = Path(ex)
        if not ex.is_absolute():
            ex = (target / ex).resolve()
        else:
            ex = ex.resolve()
        exclude_abs.add(ex)

    def _purge(curr: Path) -> None:
        for child in curr.iterdir():
            try:
                cres = child.resolve()

                # ----------------------------
                # 1) 폴더 자체가 exclude면: 통째로 건드리지 않음
                # ----------------------------
                if cres in exclude_abs and child.is_dir():
                    print(f"[SKIP] 제외 폴더 보존: {child}")
                    continue

                # ----------------------------
                # 2) 파일/링크 처리
                # ----------------------------
                if child.is_file() or child.is_symlink():
                    if cres in exclude_abs:
                        print(f"[SKIP] 제외 파일 보존: {child}")
                        continue
                    if dry_run:
                        print(f"[DRY ] 파일 삭제 예정: {child}")
                    else:
                        child.unlink()
                        print(f"[DEL ] 파일 삭제: {child}")
                    continue

                # ----------------------------
                # 3) 폴더 처리 (중요)
                #    - **폴더는 절대 삭제하지 않음**
                #    - 내부로 재귀 진입하여 파일만 정리
                # ----------------------------
                if child.is_dir():
                    _purge(child)
                    # 폴더 삭제 로직은 전부 제거 (폴더 구조는 항상 유지)
            except Exception as e:
                print(f"[ERR ] 삭제 실패: {child} -> {e}")

    _purge(target)


if __name__ == "__main__":
    # ──────────────────────────────────────────────────────────
    # 1) daily_news 폴더: keyword_news.xlsx는 보존(삭제 제외)
    #    → 상대경로 'keyword_news.xlsx'는 TARGET_DAILY_NEWS 기준으로 해석됨
    # ──────────────────────────────────────────────────────────
    EXCLUDE_IN_DAILY = [Path("keyword_news.xlsx")]

    # [변경] 2) bigkinds 전체 정리 시에도 keyword_news.xlsx 보존
    # - 여기서는 절대경로로 예외를 넣어줌
    # - TARGET_BIGKINDS 아래 어디를 돌더라도 이 파일은 삭제되지 않게 보호
    EXCLUDE_IN_BIGKINDS = [
        TARGET_DAILY_NEWS / "keyword_news.xlsx"   # [변경] 절대경로 예외 지정
    ]

    # (샘플) 예행연습: 무엇이 지워질지 로그만 확인할 때 사용
    # purge_dir(TARGET_DAILY_NEWS, dry_run=True,  exclude=EXCLUDE_IN_DAILY)
    # purge_dir(TARGET_BIGKINDS,   dry_run=True,  exclude=EXCLUDE_IN_BIGKINDS)

    # 실제 실행: **폴더는 남기고 파일/링크만 삭제**
    purge_dir(TARGET_DAILY_NEWS, dry_run=False, exclude=EXCLUDE_IN_DAILY)
    purge_dir(TARGET_BIGKINDS,   dry_run=False, exclude=EXCLUDE_IN_BIGKINDS)  # [변경]


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import os
import time
import datetime
import shutil

# 📌 다운로드 폴더 설정
download_folder = r"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds"

# 폴더가 없으면 생성
if not os.path.exists(download_folder):
    os.makedirs(download_folder)

# 🔹 오늘 날짜 가져오기 (YYYYMMDD 형식)
today_date = datetime.datetime.today().strftime("%Y%m%d")

# 🔹 크롬 옵션 설정 (자동 다운로드 활성화)
chrome_options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": download_folder,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)

# 🔹 크롬 드라이버 실행
driver = webdriver.Chrome(options=chrome_options)

# 🔹 사이트 접속
driver.get("https://www.bigkinds.or.kr/v2/mypage/myKeyword.do")

# 🔹 로그인 버튼에 마우스 오버하여 작은 팝업 표시
time.sleep(2)  # 페이지 로딩 대기
login_menu = driver.find_element(By.CSS_SELECTOR, "#header > div.header-anim > div > div > div.topRightArea > ul > li.topMembership > a")
ActionChains(driver).move_to_element(login_menu).perform()

# 🔹 작은 팝업에서 로그인 링크 클릭
time.sleep(1)  # 팝업 표시 대기
login_link = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/div/div/div[2]/ul/li[1]/div/ul/li/a')
login_link.click()

# 🔹 로그인 팝업에서 아이디와 비밀번호 입력
time.sleep(2)  # 로그인 팝업 로딩 대기
username = "amongddomong@gmail.com"
password = "^^joahane1"

# 아이디 입력
id_input = driver.find_element(By.XPATH, '//*[@id="login-user-id"]')
id_input.clear()  # 👉 기존 입력값 제거
id_input.send_keys(username)

# 비밀번호 입력
pw_input = driver.find_element(By.XPATH, '//*[@id="login-user-password"]')
pw_input.clear()  # 👉 기존 입력값 제거
pw_input.send_keys(password)

# 로그인 버튼 클릭
login_button = driver.find_element(By.XPATH, '//*[@id="login-btn"]')
login_button.click()

# 🔹 로그인 후 페이지 로딩 대기
time.sleep(5)

# 🔹 다운로드 페이지로 직접 이동
driver.get("https://www.bigkinds.or.kr/v2/mypage/myKeyword.do")

# 🔹 다운로드 버튼 클릭 (JS 강제 클릭 방식)
time.sleep(3)  # 페이지 로딩 대기
download_button = driver.find_element(By.XPATH, '//*[@id="btn-keyword-download"]')
driver.execute_script("arguments[0].click();", download_button)

# 🔹 팝업이 나타나면 "확인" 버튼 클릭
time.sleep(3)
try:
    alert = driver.switch_to.alert
    print("팝업 감지됨! 확인 버튼 클릭 중...")
    alert.accept()
    print("확인 버튼 클릭 완료!")
except:
    print("팝업을 찾을 수 없음.")

# 🔹 다운로드 완료 대기
time.sleep(50)

# 🔹 다운로드된 파일 이름 변경 (가장 최근 파일 찾기)
downloaded_files = os.listdir(download_folder)
if downloaded_files:
    downloaded_files = sorted(
        downloaded_files,
        key=lambda x: os.path.getctime(os.path.join(download_folder, x)),
        reverse=True
    )
    latest_file = os.path.join(download_folder, downloaded_files[0])
    new_filename = os.path.join(download_folder, f"{today_date}_bigkinds.xlsx")

    shutil.move(latest_file, new_filename)
    print(f"파일명이 '{new_filename}'(으)로 변경되었습니다!")

print(f"파일이 {download_folder} 폴더에 저장되었습니다!")

# 🔹 드라이버 종료
driver.quit()


import pandas as pd
import os
from datetime import datetime, timedelta

# 🔹 오늘 날짜 문자열 생성 (YYYYMMDD)
today = datetime.now()
today_str = today.strftime("%Y%m%d")

# 🔹 하루 전 날짜 문자열 생성
yesterday = today - timedelta(days=1)
yesterday_str = yesterday.strftime("%Y%m%d")

# 🔹 원본 파일 경로 (오늘 기준)
input_path = rf"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds\{today_str}_bigkinds.xlsx"

# 🔹 결과 저장 파일 경로
output_path = rf"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds\{today_str}_bigkinds_risk.xlsx"

# 🔹 엑셀 파일 로드
df = pd.read_excel(input_path)

# 🔹 '일자' 컬럼이 datetime이 아니라면 문자열로 처리
df["일자"] = df["일자"].astype(str)

# 🔹 오늘과 하루 전 날짜 포함하는 데이터만 필터링
df = df[df["일자"].isin([today_str, yesterday_str])]

# 🔹 필요한 컬럼만 선택
df = df[['일자', '제목', '언론사', '통합 분류1', 'URL']]

# 🔹 제외할 통합 분류1 앞글자 리스트
excluded_prefixes = ['IT_과학', '문화', '미분류', '스포츠']

# 🔹 제외 조건 적용
df = df[~df["통합 분류1"].str.startswith(tuple(excluded_prefixes), na=False)]

# 🔹 URL 값이 NULL인 행 제거
df = df[df["URL"].notna()]

# 🔹 결과 저장
df.to_excel(output_path, index=False)
print(f"✅ 오늘과 하루 전 날짜 기준 필터링된 파일이 저장되었습니다: {output_path}")

# 🔹 원본 파일 삭제
if os.path.exists(input_path):
    os.remove(input_path)
    print(f"🗑️ 원본 파일 삭제 완료: {input_path}")
else:
    print(f"⚠️ 원본 파일이 존재하지 않습니다: {input_path}")


import pandas as pd
import datetime as dt
import os
import re
import glob
from typing import List, Optional


def _pick_col(df: pd.DataFrame, candidates: List[str], required: bool = True) -> Optional[str]:
    """
    다양한 표기(대소문자/한영혼용/스페이스 유무)를 고려해 가장 먼저 매칭되는 컬럼명을 찾아 반환.
    required=False이면 못 찾을 경우 None 반환.
    """
    norm = {re.sub(r"\s+", "", str(c)).lower(): c for c in df.columns}
    for cand in candidates:
        key = re.sub(r"\s+", "", cand).lower()
        if key in norm:
            return norm[key]
    if required:
        raise KeyError(f"필수 컬럼을 찾을 수 없습니다. 후보: {candidates}")
    return None


def _load_keywords(keyword_path: str) -> pd.DataFrame:
    """
    keyword.xlsx 로드 및 표준 컬럼명으로 정리
    기대 컬럼(대체 허용): 
      - Scenario_ID: ['Scenario_ID','시나리오ID','시나리오','ScenarioId','SCENARIO_ID']
      - News_kor   : ['News_kor','news_kor','News_KOR','키워드','Keyword']
      - Phase      : ['Phase','phase','단계']
    """
    df = pd.read_excel(keyword_path)
    sc_col = _pick_col(df, ["Scenario_ID", "시나리오ID", "시나리오", "ScenarioId", "SCENARIO_ID"])
    kw_col = _pick_col(df, ["News_kor", "news_kor", "News_KOR", "키워드", "Keyword"])
    ph_col = _pick_col(df, ["Phase", "phase", "단계"], required=False)

    out = pd.DataFrame({
        "Scenario_ID": df[sc_col].astype(str).str.strip(),
        "News_kor": df[kw_col].astype(str).str.strip(),
        "Phase": df[ph_col].astype(str).str.strip() if ph_col else ""
    })
    # 빈 키워드 제거
    out = out[out["News_kor"].str.len() > 0].reset_index(drop=True)
    return out


def _load_latest_bigkinds(risk_dir: str) -> pd.DataFrame:
    """
    risk_dir 내 '*bigkinds_risk*.xlsx' 파일 중 **가장 최신 수정시각** 파일을 읽어 표준 컬럼명으로 반환
    기대 컬럼(대체 허용):
      - 제목   : ['제목','기사제목','title','Title']
      - 언론사 : ['언론사','언론사명','매체','publisher','Publisher','신문사']
      - URL   : ['URL','url','링크','Link']
    """
    paths = glob.glob(os.path.join(risk_dir, "*bigkinds_risk*.xlsx"))
    if not paths:
        raise FileNotFoundError(f"bigkinds_risk 파일을 찾을 수 없습니다: {risk_dir}")

    latest = max(paths, key=os.path.getmtime)  # [변경점] 가장 최근 파일 자동 선택
    df = pd.read_excel(latest)

    title_col = _pick_col(df, ["제목", "기사제목", "title", "Title"])
    media_col = _pick_col(df, ["언론사", "언론사명", "매체", "publisher", "Publisher", "신문사"])
    url_col   = _pick_col(df, ["URL", "url", "링크", "Link"])

    out = pd.DataFrame({
        "제목": df[title_col].astype(str).str.strip(),
        "언론사": df[media_col].astype(str).str.strip(),
        "URL": df[url_col].astype(str).str.strip()
    }).dropna(subset=["제목"]).reset_index(drop=True)

    # 중복 제목 제거(선행 데이터 우선)
    out = out.drop_duplicates(subset=["제목"], keep="first").reset_index(drop=True)
    return out


def _match_first(corpus: pd.DataFrame, query: str, used_titles: set) -> Optional[pd.Series]:
    """
    News_kor 문자열을 ',' 기준 **AND 조건**으로 분해하여 기사 제목에서 모두 매칭되는 첫 건 반환.
    이미 선택된 제목(used_titles)은 제외.
    """
    # '키워드1,키워드2' -> ['키워드1','키워드2']
    terms = [t.strip() for t in str(query).split(",") if t.strip()]
    mask = pd.Series(True, index=corpus.index)
    for t in terms:
        # 대소문자 무시, 정규식 메타문자 이스케이프
        mask &= corpus["제목"].str.contains(re.escape(t), case=False, na=False)

    candidates = corpus[mask & ~corpus["제목"].isin(used_titles)]
    if candidates.empty:
        return None
    # 첫 건 선택(가장 단순하고 예측가능)  [변경점] 우선순위언론사 규칙 제거
    return candidates.iloc[0]


def extract_daily_news_from_keyword_file(
    keyword_path: str,
    output_dir: str,
    risk_dir: str = r"C:\Users\amongpapa\uto\risk_report\bigkinds"
):
    """
    [변경점] 요구사항에 맞춘 새로운 메인 함수
      - 입력: keyword.xlsx (Scenario_ID, News_kor, Phase)
      - 빅카인즈: risk_dir의 최신 '*bigkinds_risk*.xlsx'
      - 매칭 로직: News_kor를 ','로 분해한 **AND 조건**으로 제목 검색, 키워드당 1건 선정(중복 제목은 스킵)
      - 출력 컬럼: Scenario_ID, News_kor, Phase, 제목, 언론사, URL
      - 파일명: YYYYMMDD_news.xlsx
    """
    # 1) 키워드 로드
    df_kw = _load_keywords(keyword_path)

    # 2) 최신 bigkinds 데이터 로드
    df_risk = _load_latest_bigkinds(risk_dir)

    # 3) 매칭 수행
    used_titles = set()
    rows = []
    not_found = []

    for _, r in df_kw.iterrows():
        scenario_id = r["Scenario_ID"]
        news_kor = r["News_kor"]
        phase = r["Phase"]

        hit = _match_first(df_risk, news_kor, used_titles)
        if hit is None:
            not_found.append((scenario_id, news_kor))
            continue

        used_titles.add(hit["제목"])
        rows.append({
            "Scenario_ID": scenario_id,
            "News_kor": news_kor,
            "Phase": phase,
            "제목": hit["제목"],
            "언론사": hit["언론사"],
            "URL": hit["URL"],
        })

    # 4) 저장
    os.makedirs(output_dir, exist_ok=True)
    date_str = dt.datetime.now().strftime("%Y%m%d")
    out_file = os.path.join(output_dir, f"keyword_news_bf.xlsx")
    pd.DataFrame(rows).to_excel(out_file, index=False)

    print(f"[완료] 데일리 기사 파일: {out_file}")
    print(f" - 매칭 성공 {len(rows)}건 / 실패 {len(not_found)}건")
    if not_found:
        print(" - 매칭 실패 리스트(Scenario_ID, News_kor):")
        for sc, kw in not_found[:20]:  # 너무 길면 앞 20개만
            print(f"   · {sc} | {kw}")


if __name__ == "__main__":
    # ⚠️ 경로는 r"..." raw string으로 입력하세요 (윈도우 백슬래시 이스케이프 방지)
    keyword_path = r"C:\Users\amongpapa\chartup\go_scen\data\news\keyword.xlsx"   # [변경점]
    output_dir   = r"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds\daily_news"                          # 필요에 맞게 조정
    risk_dir     = r"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds"                 # 필요 시 조정
    extract_daily_news_from_keyword_file(keyword_path, output_dir, risk_dir)


# -*- coding: utf-8 -*-
r"""
목적(업데이트 버전 - 운영용)
- (소스) daily_news 폴더의 'keyword_news_bf.xlsx' 만 처리
- 각 행의 URL 페이지를 Selenium으로 열어 body.innerText 추출 → GPT 요약 생성
- (타깃) 'keyword_news.xlsx'는 저장용: 기존 내용은 유지하면서
  - 처리된 행들을 '그대로' 아래로 이어서 추가(append)
  - 전체 행 수가 max_rows(기본 30)를 넘으면, '먼저 들어온 행'부터 제거(FIFO)하여 최근 max_rows행 유지

특징
- news 컬럼 비어 있는 행만 처리(FILL_ONLY_EMPTY=True)
- 특정 언론사에서 timeout 나더라도 그 행만 스킵하고 다음 행 처리
- Windows 경로는 pathlib.Path(r"...") 사용
"""

import re
import time
from typing import List, Optional
from pathlib import Path

import pandas as pd

# ✅ Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException, TimeoutException, NoSuchElementException

# =========================
# 경로 및 옵션
# =========================
DIR_PATH = Path(r"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds\daily_news")  # 작업 폴더
KEY_PATH = Path(r"C:\Users\amongpapa\lm\keys\open.txt")  # OpenAI 키 파일
SOURCE_BF = DIR_PATH / "keyword_news_bf.xlsx"  # 소스(매일 갱신)
TARGET_NEWS = DIR_PATH / "keyword_news.xlsx"   # 타깃(저장/롤링)

# 🔁 수정: 최대 행 수 20 → 30 (최근 30개만 유지)
MAX_TARGET_ROWS = 30  # 타깃 최대 행 수(넘치면 FIFO, 최근 30개 유지)

# ✅ daily_news 폴더가 없으면 자동 생성
DIR_PATH.mkdir(parents=True, exist_ok=True)

# 팀장님 환경
CHROME_PATH = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")
CHROME_PROFILE = "Profile 1"

MODEL_ID = "gpt-4o-mini"
FILL_ONLY_EMPTY = True
TARGET_NEWS_COL_INDEX = 6  # G열(0-based=6)
PAGE_LOAD_WAIT = 8.0       # 페이지 로딩 대기(초)
REQUEST_INTERVAL_SEC = 0.4 # GPT 호출 간격

# =========================
# OpenAI 신/구 SDK 자동 호환
# =========================
def load_api_key(path: Path) -> str:
    """open.txt에서 OpenAI API 키를 한 줄 읽어오는 함수."""
    key = path.read_text(encoding="utf-8").strip()
    if not key:
        raise ValueError("API 키가 비어 있습니다.")
    return key


class ChatClient:
    """신/구 OpenAI SDK 자동 감지 래퍼(.chat -> str)."""
    def __init__(self, api_key: str):
        self.mode = None
        try:
            from openai import OpenAI  # 신버전
            self._client = OpenAI(api_key=api_key)
            self.mode = "new"
        except Exception:
            import openai  # 구버전
            openai.api_key = api_key
            self._client = openai
            self.mode = "old"

    def chat(self, messages: List[dict], model: str, temperature: float = 0.2) -> str:
        """messages 리스트를 받아 GPT로부터 문자열 답변만 반환하는 공통 인터페이스."""
        if self.mode == "new":
            res = self._client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res.choices[0].message.content.strip()
        else:
            res = self._client.ChatCompletion.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res["choices"][0]["message"]["content"].strip()


# =========================
# 텍스트 정리 & 프롬프트 구성
# =========================
PROMPT_TEMPLATE = (
    "기사 요약 – 서술형 보고서\n"
    "[역할] 당신은 은행 리스크관리 보고서 에디터다.\n"
    "[목표] 여러 기사 원문에서 광고/홍보/무관/중복 내용을 제거하고 핵심 사실만을 바탕으로 한국어 서술형 보고서를 작성한다. "
    "총 길이 {2000}자 이내.\n"
    "[입력] {기사 원문들}\n"
    "[출력 지침 – 서술형 보고서]\n"
    "- 제목: 한 줄\n"
    "- 요약: 3문장\n"
    "- 본문: 연대기 순(사실→원인→영향→전망)으로 3–6단락. 숫자·기관명·지명 정확히 기술.\n"
    "- 핵심 수치: 본문에 자연스럽게 녹여 쓰되 최초 등장 문장 끝에 근거 [s#] 표기(최대 5개).\n"
    "- 자산/섹터/티커: 필요한 경우 괄호로 간단 표기(예: 원/달러 환율(USDKRW), 삼성전자(005930.KS)).\n"
    "- 시나리오 연결: 문단 말미에 1–2문장으로 ‘주요 동인·지표·트리거·방향성·예상 시계’를 통합 서술.\n"
    "- 신뢰도: 마지막 문장에 (신뢰도 0.0–1.0; 근거 요약) 형식으로 표기.\n"
    "- 중복/제거: 삭제·통합한 기사 제목(또는 요지)만 마지막 줄에 ‘제거: …’로 나열.\n"
    "[금지] 광고문구, 과장/추측, 불필요 텍스트, 표/불릿/JSON 출력.\n"
    "[실행] 위 지침을 적용해 결과만 출력한다.\n"
)


def clean_text(text: str) -> str:
    """
    본문 텍스트 정제:
    - 개행/공백 정리
    - 쿠키/저작권/구독/광고 관련 문구 제거
    """
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)

    drop_patterns = [
        r"쿠키(를|에) 사용",
        r"이용약관",
        r"개인정보",
        r"구독",
        r"광고문의",
        r"무단전재",
        r"저작권",
    ]

    cleaned_lines = []
    for ln in text.splitlines():
        if any(re.search(pat, ln, re.IGNORECASE) for pat in drop_patterns):
            continue
        cleaned_lines.append(ln.strip())
    return "\n".join(cleaned_lines).strip()


def build_prompt(article_text: str, max_chars: int = 2000) -> str:
    """
    기사 텍스트를 PROMPT_TEMPLATE에 채워 넣어 GPT용 프롬프트 생성.
    - 너무 긴 텍스트는 8,000자까지만 사용
    """
    article_text = article_text[:8000]
    return (
        PROMPT_TEMPLATE
        .replace("{2000}", str(max_chars))
        .replace("{기사 원문들}", article_text)
    )


def summarize_article(client: ChatClient, article_text: str, model: str = MODEL_ID) -> str:
    """
    정제된 기사 텍스트 → GPT 요약.
    - 시스템 메시지로 '리스크관리 보조원' 역할 지정
    - 출력은 반드시 한국어 서술형
    """
    sys_msg = (
        "너는 은행 리스크관리 보조원이며, 출력은 반드시 한국어로 작성한다. "
        "광고/구독/저작권/추천기사/댓글 등 불필요 텍스트는 제거하고, "
        "요구된 필드만 간결하게 작성한다."
    )

    messages = [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": build_prompt(article_text, max_chars=2000)},
    ]
    return client.chat(messages, model=model, temperature=0.2)


# =========================
# Selenium 드라이버 구성/해제
# =========================
def build_driver(download_dir: Optional[Path] = None) -> webdriver.Chrome:
    """
    Selenium Chrome 드라이버 생성.
    - 크롬 프로필 재사용(회사 프록시/보안 설정 그대로 활용)
    - page_load_timeout 기본 45초
    """
    chrome_options = webdriver.ChromeOptions()

    # 팀장님 PC에 설치된 Chrome 경로 지정
    if CHROME_PATH.exists():
        chrome_options.binary_location = str(CHROME_PATH)

    # 지정한 사용자 프로필 사용
    chrome_options.add_argument(f'--profile-directory={CHROME_PROFILE}')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-notifications')
    chrome_options.add_argument('--start-maximized')

    # (선택) 다운로드 경로 지정
    if download_dir:
        prefs = {
            "download.default_directory": str(download_dir),
            "download.prompt_for_download": False,
            "directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(45)  # 🔁 30초 → 45초로 여유
    return driver


def safe_quit(driver: webdriver.Chrome):
    """드라이버 종료 시 예외 무시하고 안전하게 종료."""
    try:
        driver.quit()
    except Exception:
        pass


# =========================
# 엑셀 유틸
# =========================
def ensure_news_column_at_G(df: pd.DataFrame) -> pd.DataFrame:
    """
    'news' 칼럼을 G열(index=6)에 고정:
    - 없으면 생성하여 G열에 삽입
    - 있으면 해당 칼럼을 G열 위치로 이동
    """
    cols = list(df.columns)

    if "news" not in cols:
        insert_at = min(TARGET_NEWS_COL_INDEX, len(cols))
        df.insert(insert_at, "news", pd.NA)
    else:
        current_idx = cols.index("news")
        if current_idx != TARGET_NEWS_COL_INDEX:
            cols.pop(current_idx)
            insert_at = min(TARGET_NEWS_COL_INDEX, len(cols))
            cols.insert(insert_at, "news")
            df = df.reindex(columns=cols)

    return df


def should_fill(val) -> bool:
    """
    news를 채울지 여부(FILL_ONLY_EMPTY 정책).
    - FILL_ONLY_EMPTY=True: 비어 있을 때만 True
    - pandas의 <NA>, NaN 등도 비어있는 것으로 처리
    """
    if not FILL_ONLY_EMPTY:
        # 항상 덮어쓰는 모드일 때는 그냥 True
        return True

    # 어떤 타입이든 NaN/<NA>/None이면 True
    if pd.isna(val):
        return True

    # 문자열인데 공백뿐인 경우도 비어있는 것으로 간주
    if isinstance(val, str) and not val.strip():
        return True

    return False


# =========================
# 페이지 텍스트 추출 (Selenium)
# =========================
def get_visible_text(driver: webdriver.Chrome, url: str) -> str:
    """
    주어진 URL을 열고, document.body.innerText를 가져옴.
    - PAGE_LOAD_WAIT 동안 추가 대기
    - timeout/기타 예외 시 빈 문자열 반환
    """
    try:
        print(f"[INFO] 페이지 열기: {url}")
        driver.get(url)
        time.sleep(PAGE_LOAD_WAIT)

        try:
            text = driver.execute_script(
                "return document.body ? document.body.innerText : '';"
            )
        except Exception:
            try:
                text = driver.find_element(By.TAG_NAME, "body").text
            except NoSuchElementException:
                text = ""

        text = text or ""
        print(f"[DEBUG] 추출된 텍스트 길이: {len(text)}")
        return text

    except (WebDriverException, TimeoutException) as e:
        print(f"[ERR] 페이지 로딩 실패: {e}")
        return ""


# =========================
# 타깃 파일(keyword_news.xlsx) Append + FIFO(최근 30행 유지)
# =========================
def append_rows_to_keyword_news(
    rows_df: pd.DataFrame,
    target_path: Path,
    max_rows: int = MAX_TARGET_ROWS
) -> None:
    """
    - 기존 target을 읽어오고(rows_old), rows_df를 아래로 이어붙임
    - 전체 행이 max_rows를 초과하면 '먼저 들어온 행'부터 제거하여 최근 max_rows행 유지
    """
    if target_path.exists():
        try:
            rows_old = pd.read_excel(target_path, engine="openpyxl")
            print(f"[INFO] 기존 타깃 행 수: {len(rows_old)}")
        except Exception as e:
            print(f"[WARN] 타깃 파일 로딩 실패, 이번 실행부터 새로 시작: {e}")
            rows_old = pd.DataFrame()
    else:
        print("[INFO] 타깃 파일 최초 생성 상태")
        rows_old = pd.DataFrame()

    rows_old = ensure_news_column_at_G(rows_old)
    rows_df = ensure_news_column_at_G(rows_df.copy())

    # 기존 + 신규를 그대로 이어 붙임 (keyword_news.xlsx 마지막부터 이어서)
    out_df = pd.concat([rows_old, rows_df], ignore_index=True)

    # FIFO: 최대 행 수 유지 (앞에서 잘라내고 최신 max_rows만 남김)
    if len(out_df) > max_rows:
        out_df = out_df.iloc[-max_rows:].reset_index(drop=True)

    out_df = ensure_news_column_at_G(out_df)
    out_df.to_excel(target_path, index=False, engine="openpyxl")
    print(f"[OK] 저장 완료(target): {target_path.name} (총 {len(out_df)}행, max={max_rows})")


# =========================
# 소스 파일 처리 + '이번에 처리된 행' 반환
# =========================
def process_source_file_and_collect(
    xlsx_path: Path,
    client: ChatClient,
    driver: webdriver.Chrome
) -> pd.DataFrame:
    """
    - (소스) keyword_news_bf.xlsx 한 개를 처리
    - 이번 실행에서 news가 새로 채워진 행만 모아 DataFrame으로 반환(타깃 Append 용)
    - 소스 파일에도 news 반영하여 저장
    """
    try:
        df = pd.read_excel(xlsx_path, engine="openpyxl")
    except Exception as e:
        print(f"[ERR] 엑셀 로딩 실패: {xlsx_path.name} -> {e}")
        return pd.DataFrame()

    if "URL" not in df.columns:
        print(f"[SKIP] 'URL' 칼럼 없음: {xlsx_path.name}")
        return pd.DataFrame()

    df = ensure_news_column_at_G(df)

    updated_indices: List[int] = []

    for idx, row in df.iterrows():
        url = str(row.get("URL", "")).strip()
        news_val = row.get("news", None)
        fill_flag = should_fill(news_val)

        print(f"\n[DEBUG] row={idx}, fill?={fill_flag}, news={repr(news_val)}")
        print(f"[DEBUG] URL={url}")

        if not url.startswith("http"):
            print("[SKIP] http/https 아님")
            continue

        if not fill_flag:
            print("[SKIP] news가 이미 채워져 있음(덮어쓰기 모드 아님)")
            continue

        # 1) 페이지 텍스트 추출
        raw = get_visible_text(driver, url)
        if not raw or len(raw) < 30:
            print(f"[WARN] 텍스트가 너무 짧음(len={len(raw)})")
            continue

        # 2) 텍스트 정제
        article_text = clean_text(raw)
        print(f"[DEBUG] 정제 후 텍스트 길이: {len(article_text)}")

        # 3) GPT 요약
        try:
            summary = summarize_article(client, article_text, model=MODEL_ID)
        except Exception as e:
            print(f"[ERR] GPT 요약 실패 row={idx} -> {e}")
            continue

        # 4) 소스 df에 news 반영
        df.at[idx, "news"] = summary
        updated_indices.append(idx)
        print(f"[OK] row={idx} 요약 완료")

        # 5) API 호출 간 간격
        time.sleep(REQUEST_INTERVAL_SEC)

    # 소스 파일 저장
    try:
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        print(f"\n[OK] 저장 완료(source): {xlsx_path.name} (updated={len(updated_indices)})")
    except Exception as e:
        print(f"[ERR] 소스 저장 실패: {xlsx_path.name} -> {e}")

    if updated_indices:
        # 이번에 실제로 새로 요약된 행만 반환 (타깃 append 용)
        return df.loc[updated_indices].copy()

    return pd.DataFrame()


# =========================
# 실행부
# =========================
def main():
    # 1) API 키 로딩
    api_key = load_api_key(KEY_PATH)
    client = ChatClient(api_key=api_key)

    # 2) Selenium 드라이버 생성
    driver = build_driver(download_dir=None)

    try:
        bf_path = SOURCE_BF

        if not bf_path.exists():
            print(f"[ERR] 소스 파일 없음: {bf_path}")
            return

        print(f"[INFO] 소스 파일: {bf_path.name}")
        print(f"[INFO] 타깃 파일: {TARGET_NEWS.name} (최대 {MAX_TARGET_ROWS}행, FIFO)")

        # 3) 소스 파일 처리 (keyword_news_bf.xlsx의 news 빈 칸 채우기)
        processed_rows = process_source_file_and_collect(bf_path, client, driver)

        # 4) 타깃 파일 Append + FIFO 유지
        #    - keyword_news.xlsx 기존 행 뒤에 processed_rows(이번에 새로 요약된 행) 이어 붙이고
        #    - 상위부터 삭제해서 최근 기준 30개만 남김
        if not processed_rows.empty:
            append_rows_to_keyword_news(
                processed_rows,
                TARGET_NEWS,
                max_rows=MAX_TARGET_ROWS
            )
        else:
            print("[INFO] 이번 실행에서 새로 처리된 행이 없습니다(요약 대상 없음).")

    finally:
        safe_quit(driver)


if __name__ == "__main__":
    main()


