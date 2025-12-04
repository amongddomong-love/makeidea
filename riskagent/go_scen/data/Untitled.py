# -*- coding: utf-8 -*-
"""
IND001 ~ IND151 엑셀 파일에 대해:
1) 시계열(날짜, 값) 자동 탐지
2) 최근 평균/표준편차 기반 임계치 계산
3) 마지막 값이 임계치를 기준으로
      - 정상       → C1 = 'G'
      - 주의/경고/심각 → C1 = 'Y'  (주의 이상은 모두 Y)
4) D1에 60일 평균(mu), E1에 60일 표준편차(sd=σ) 기록
5) 실행 요약 CSV 저장

설치:
    pip install pandas openpyxl numpy

사용:
    python script.py
"""

import os
import re
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Optional, Tuple

# ================= 팀장님 환경 설정 =================
BASE_DIR = r"C:\Users\amongpapa\chartup\go_scen\data\set"  # 엑셀 폴더 경로 (raw string 사용)
FILE_PREFIX = "IND"                                       # 파일 접두사
FILE_RANGE = range(1, 152)                                # 001~151

SHEET_NAME = 0                  # 첫 번째 시트(정수 또는 시트명)
WINDOW = 60                     # 롤링 윈도우(일수, 평균/표준편차 계산용)
K = 2.0                         # (현재는 z-score 참고용, 임계치에는 사용하지 않음)
MIN_PERIODS = max(30, WINDOW // 2)  # 최소 계산 데이터 길이(짧은 데이터 보호)

# 컬럼 이름 힌트(우선 매칭)
DATE_COL_HINTS = ["date", "날짜", "일자", "time", "일시"]
VALUE_COL_HINTS = ["close", "price", "value", "index", "지수", "종가", "가격", "값", "수치", "PX_LAST"]
# ====================================================

# ✅ [추가] 지표별 고정 임계치 설정 (방식 A: 절대값 기준)
# - 여기 있는 ID는 고정 임계치를 사용
# - 여기에 없는 ID는 자동 계산 임계치(평균 × 1.1 / 1.2 / 1.3)를 사용
CUSTOM_THRESHOLDS = {
    # USD/KRW 환율 (IND071): 값이 커질수록 리스크↑
    "IND071": {
        "direction": "up",   # up: 값이 커질수록 위험, down: 값이 작아질수록 위험
        "yellow": 1400.0,    # 주의
        "orange": 1600.0,    # 경고
        "red": 1800.0        # 심각
    },
    # 필요하면 코스피/미국 금리 등 추가 가능
    # "IND001": {
    #     "direction": "down",
    #     "yellow": 2400.0,
    #     "orange": 2200.0,
    #     "red": 2000.0
    # },
    # "IND050": {
    #     "direction": "up",
    #     "yellow": 4.5,
    #     "orange": 5.0,
    #     "red": 5.5
    # },
}


# ---------------- 도우미: 중복 컬럼명 유일화 ----------------
def _make_unique(names):
    """
    동일한 컬럼명이 반복될 경우 .1, .2 접미사를 부여해 유일화
    예) ['nan','nan'] -> ['nan','nan.1']
    """
    seen = {}
    out = []
    for n in names:
        key = str(n)
        if key not in seen:
            seen[key] = 0
            out.append(key)
        else:
            seen[key] += 1
            out.append(f"{key}.{seen[key]}")
    return out


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    공백 정리 + 유일화
    """
    cols = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    df.columns = _make_unique(cols)
    return df


# ---------------- 유틸: 숫자 문자열 → float (항상 Series 반환) ----------------
def _as_numeric_series(s) -> pd.Series:
    """
    어떤 입력(s: Series/DataFrame/ndarray/리스트)이 와도 **항상 pandas.Series**로 변환 후
    안전하게 숫자로 변환합니다.
    - 공백/대시류("", "-", "—", "_") → NaN
    - 괄호 음수 "(123)" → -123
    - 콤마 제거 "1,234.56" → 1234.56
    - 퍼센트 "5.2%" → 0.052
    """
    if isinstance(s, pd.DataFrame):
        # 숫자 변환률이 가장 높은 컬럼 하나 자동 선택
        best_col = None
        best_ratio = -1.0
        for col in s.columns:
            x = pd.to_numeric(
                pd.Series(s[col]).astype("string")
                .str.replace(",", "", regex=False)
                .str.replace("%", "", regex=False),
                errors="coerce",
            )
            ratio = x.notna().mean()
            if ratio > best_ratio:
                best_ratio, best_col = ratio, col
        s = s[best_col]

    if not isinstance(s, pd.Series):
        s = pd.Series(s)

    ss = s.astype("string").str.strip()
    ss = ss.replace(
        {
            "": pd.NA,
            "-": pd.NA,
            "—": pd.NA,
            "_": pd.NA,
            "nan": pd.NA,
            "NaN": pd.NA,
            "None": pd.NA,
        }
    )

    # 괄호 음수 처리
    neg_mask = ss.str.match(r"^\(.*\)$", na=False)
    ss2 = ss.str.replace(r"^\((.*)\)$", r"\1", regex=True)  # '(123)' -> '123'

    # 콤마/퍼센트 제거
    ss2 = ss2.str.replace(",", "", regex=False)
    pct_mask = ss2.str.endswith("%", na=False)
    ss2 = ss2.str.replace("%", "", regex=False)

    # 숫자 변환
    num = pd.to_numeric(ss2, errors="coerce")

    # 괄호 음수 적용
    if neg_mask.any():
        num.loc[neg_mask & num.notna()] = -num.loc[neg_mask & num.notna()].abs()

    # 퍼센트 → /100
    if pct_mask.any():
        num.loc[pct_mask & num.notna()] = num.loc[pct_mask & num.notna()] / 100.0

    return num.astype("float64")


# ---------------- 스코어링 헬퍼 ----------------
def _numeric_score(df: pd.DataFrame, lookahead: int = 5) -> int:
    """
    상위 lookahead행에서 숫자로 변환 가능한 열 개수 스코어링.
    """
    head = df.head(lookahead)
    cnt = 0
    for c in head.columns:
        s = head[c]
        s_num = _as_numeric_series(s)
        numlike = s_num.notna().mean()
        if numlike >= 0.5:
            cnt += 1
    return cnt


def _likely_good(df: pd.DataFrame) -> bool:
    """현재 형태가 '헤더-데이터'로 적절한지 간단 스코어."""
    return _numeric_score(df, lookahead=5) >= 2  # 숫자형 열 2개 이상이면 OK


# ---------------- 헤더 자동 감지 ----------------
def detect_header_and_read(path, sheet_name=0, max_scan=10) -> pd.DataFrame:
    """
    헤더/데이터 시작행이 불명확한 엑셀을 위한 로더.
    1) 일반 로드 후 간이 점검, 괜찮으면 그대로 사용 (컬럼 유일화 적용)
    2) 상단 max_scan행을 후보 헤더로 가정해 스코어링 후 베스트 헤더 채택
    """
    # 1) 일반 로드 시도
    df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=0)
    if df0.shape[0] == 0:
        df0 = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
    if _likely_good(df0):
        return _clean_columns(df0)

    # 2) 후보 헤더 스캔
    raw = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
    nrows = min(max_scan, len(raw))
    best_score, best_row = -1e9, 0
    for hdr in range(nrows):
        tmp = raw.copy()
        tmp.columns = _make_unique(tmp.iloc[hdr].astype(str).str.strip())
        tmp = tmp.iloc[hdr + 1:].reset_index(drop=True)

        score_num = _numeric_score(tmp, lookahead=5)
        penalty = sum(
            1
            for name in tmp.columns
            if str(name).lower() in ("nan", "nat", "", "none")
        )
        score = score_num - 0.5 * penalty  # 나쁜 헤더 패널티

        if score > best_score:
            best_score, best_row = score, hdr

    raw.columns = _make_unique(raw.iloc[best_row].astype(str).str.strip())
    df = raw.iloc[best_row + 1:].reset_index(drop=True)
    return _clean_columns(df)


# ---------------- 날짜/값 컬럼 자동 선택 ----------------
def _pick_date_col(df: pd.DataFrame) -> Optional[str]:
    # 힌트 우선
    for hint in DATE_COL_HINTS:
        for c in df.columns:
            if hint.lower() in str(c).lower():
                return c
    # dtype/변환률 기반
    dt_cols = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
    if dt_cols:
        return dt_cols[0]
    best_col, best_ratio = None, 0.0
    for c in df.columns:
        try:
            converted = pd.to_datetime(df[c], errors="coerce", infer_datetime_format=True)
            ratio = converted.notna().mean()
            if ratio > best_ratio and ratio >= 0.7:
                best_col, best_ratio = c, ratio
        except Exception:
            pass
    return best_col


def _pick_value_col(df: pd.DataFrame, exclude_cols: list) -> Optional[str]:
    """
    값 컬럼 자동 선택(힌트 → 통계 스코어).
    """
    # 힌트 우선
    for hint in VALUE_COL_HINTS:
        for c in df.columns:
            if c in exclude_cols:
                continue
            if hint.lower() in str(c).lower():
                return c

    # 숫자형 후보 스코어링(결측↓, 분산>0, 샘플 수 충분)
    candidates = []
    for c in df.columns:
        if c in exclude_cols:
            continue
        series = _as_numeric_series(df[c])
        na_ratio = pd.isna(series).mean()
        var = np.nanvar(series.astype(float))
        if (~pd.isna(series)).sum() >= MIN_PERIODS and var > 0:
            candidates.append((c, na_ratio, var))

    if not candidates:
        # MIN_PERIODS 미만이어도 최선의 숫자형 컬럼을 fallback
        fallback = []
        for c in df.columns:
            if c in exclude_cols:
                continue
            series = _as_numeric_series(df[c])
            var = np.nanvar(series.astype(float))
            if var > 0:
                na_ratio = pd.isna(series).mean()
                fallback.append((c, na_ratio, var))
        if not fallback:
            return None
        fallback.sort(key=lambda x: (x[1], -x[2]))
        return fallback[0][0]

    candidates.sort(key=lambda x: (x[1], -x[2]))
    return candidates[0][0]


# ---------------- 로딩/계산 ----------------
def _load_timeseries(path: str, sheet_name=0) -> Tuple[pd.Series, pd.DataFrame]:
    """
    엑셀에서 시계열(날짜, 값)을 추출해 Series 반환.
    - 날짜 컬럼이 있으면 DatetimeIndex 정렬
    - 없으면 단순 순번 인덱스
    """
    df = detect_header_and_read(path, sheet_name=sheet_name)

    date_col = _pick_date_col(df)
    value_col = _pick_value_col(df, exclude_cols=[date_col] if date_col else [])

    if value_col is None:
        raise ValueError("숫자형 시계열 컬럼을 찾지 못했습니다. (헤더/형식 확인 필요)")

    if date_col is not None:
        dt = pd.to_datetime(df[date_col], errors="coerce", infer_datetime_format=True)
        df = df.loc[dt.notna()].copy()
        df.index = pd.to_datetime(df[date_col], errors="coerce")
        df = df.sort_index()
    else:
        df = df.reset_index(drop=True)

    s = _as_numeric_series(df[value_col]).astype(float).dropna()

    # 같은 날짜 중복 → 평균
    if isinstance(df.index, pd.DatetimeIndex) and s.index.has_duplicates:
        s = s.groupby(level=0).mean()

    return s.sort_index(), df


# ---------------- 임계치 기반 플래그 계산 ----------------
def compute_threshold_flag(
    s: pd.Series,
    indicator_id: Optional[str] = None,
    window: int = 60,
    k: float = 2.0,
) -> Tuple[str, dict]:
    """
    1) 60일 롤링 평균(mu), 표준편차(sd) 계산
    2) 임계치 결정
       - CUSTOM_THRESHOLDS에 등록된 ID: 고정 임계치 사용
       - 그 외: mu×1.1, 1.2, 1.3 자동 임계치 사용
    3) 마지막 값의 레벨 판단
       - level ∈ {normal, yellow, orange, red}
       - 팀장님 요청: 주의 이상(yellow, orange, red)은 모두 'Y', 그 외는 'G'
    반환:
      flag: 'G' 또는 'Y'
      info: 각종 참고 정보(dict)
    """
    if len(s) < MIN_PERIODS:
        # 데이터가 너무 짧으면 보수적으로 G 처리
        return "G", {"reason": "too_short", "len": len(s)}

    roll_mean = s.rolling(window=window, min_periods=MIN_PERIODS).mean()
    roll_std = s.rolling(window=window, min_periods=MIN_PERIODS).std(ddof=0)

    last_val = s.iloc[-1]
    mu = roll_mean.iloc[-1]
    sd = roll_std.iloc[-1]

    if pd.isna(mu) or pd.isna(sd) or sd == 0:
        return "G", {
            "reason": "nan_or_zero_std",
            "mu": mu,
            "sd": sd,
            "last": last_val,
        }

    # z-score는 참고용으로만 계산
    z = (last_val - mu) / sd

    # 1) 임계치 결정 (고정/자동)
    ind_id = (indicator_id or "").upper()
    cfg = CUSTOM_THRESHOLDS.get(ind_id)

    if cfg is not None:
        # ✅ 고정 임계치 사용
        direction = cfg.get("direction", "up")
        thr_yellow = cfg["yellow"]
        thr_orange = cfg["orange"]
        thr_red = cfg["red"]
        reason = "custom_threshold"
    else:
        # ✅ 자동 임계치 (평균 × 비율)
        direction = "up"  # 기본: 값이 올라갈수록 위험한 지표로 가정
        thr_yellow = mu * 1.1
        thr_orange = mu * 1.2
        thr_red = mu * 1.3
        reason = "auto_threshold"

    # 2) 레벨 판정
    if direction == "up":
        if last_val >= thr_red:
            level = "red"
        elif last_val >= thr_orange:
            level = "orange"
        elif last_val >= thr_yellow:
            level = "yellow"
        else:
            level = "normal"
    else:  # direction == "down": 값이 작아질수록 위험
        if last_val <= thr_red:
            level = "red"
        elif last_val <= thr_orange:
            level = "orange"
        elif last_val <= thr_yellow:
            level = "yellow"
        else:
            level = "normal"

    # 3) 팀장님 요청: "주의 이상은 Y" → normal만 G, 나머지는 전부 Y
    if level == "normal":
        flag = "G"
    else:
        flag = "Y"

    info = {
        "last": last_val,
        "mu": mu,
        "sd": sd,
        "zscore": z,
        "thr_yellow": thr_yellow,
        "thr_orange": thr_orange,
        "thr_red": thr_red,
        "level": level,
        "reason": reason,
    }

    # 예전 구조 호환용 필드(upper/lower)는 None으로 둠
    info["upper"] = None
    info["lower"] = None

    return flag, info


# ---------------- 엑셀 기록 ----------------
def write_results_to_excel(
    path: str,
    flag: str,
    mu: Optional[float],
    sd: Optional[float],
    sheet_name=0,
):
    """
    엑셀 파일의 C1/D1/E1 업데이트:
      - C1: G 또는 Y (주의 이상은 모두 Y)
      - D1: 60일 롤링 평균(mu)
      - E1: 60일 롤링 표준편차(sd = σ)
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

    ws["C1"] = flag

    def _num_or_none(x):
        try:
            if x is None:
                return None
            if isinstance(x, (int, float)) and not (
                isinstance(x, float) and np.isnan(x)
            ):
                return float(x)
            return None
        except Exception:
            return None

    ws["D1"] = _num_or_none(mu)  # 60일 평균
    ws["E1"] = _num_or_none(sd)  # 60일 표준편차(σ)

    wb.save(path)


# ---------------- 메인 드라이버 ----------------
def main():
    results = []
    for i in FILE_RANGE:
        fname = f"{FILE_PREFIX}{i:03d}.xlsx"
        fpath = os.path.join(BASE_DIR, fname)
        if not os.path.exists(fpath):
            results.append((fname, "missing", None))
            print(f"[SKIP] {fname} (file not found)")
            continue

        try:
            s, _df = _load_timeseries(fpath, sheet_name=SHEET_NAME)

            # ✅ [변경] 지표 ID를 넘겨서 임계치 기반 플래그 계산
            indicator_id = f"{FILE_PREFIX}{i:03d}"
            flag, info = compute_threshold_flag(
                s, indicator_id=indicator_id, window=WINDOW, k=K
            )

            mu = info.get("mu") if isinstance(info, dict) else None
            sd = info.get("sd") if isinstance(info, dict) else None
            write_results_to_excel(fpath, flag, mu, sd, sheet_name=SHEET_NAME)

            ztxt = (
                f"{info.get('zscore'):.2f}"
                if isinstance(info, dict)
                and "zscore" in info
                and pd.notna(info.get("zscore"))
                else "n/a"
            )
            mutxt = (
                f"{mu:.6g}"
                if isinstance(mu, (int, float)) and not pd.isna(mu)
                else "n/a"
            )
            sdtxt = (
                f"{sd:.6g}"
                if isinstance(sd, (int, float)) and not pd.isna(sd)
                else "n/a"
            )

            level = info.get("level") if isinstance(info, dict) else "n/a"

            print(
                f"[OK] {fname} -> C1='{flag}', level={level}, "
                f"D1(mu)={mutxt}, E1(sd)={sdtxt} (z={ztxt})"
            )
            results.append((fname, flag, info))
        except Exception as e:
            err_msg = str(e)
            print(f"[ERR] {fname}: {err_msg}")
            # 간단 프로브: 컬럼/타입 힌트 출력 (헤더/형식 추적)
            try:
                df_probe = detect_header_and_read(fpath, sheet_name=SHEET_NAME)
                print(f"  -> columns: {list(df_probe.columns)}")
                sample_info = []
                for c in df_probe.columns[:10]:
                    sample_info.append(f"{c}:{str(df_probe[c].dtype)}")
                print("  -> dtypes(head):", ", ".join(sample_info))
            except Exception as e2:
                print(f"  -> probe failed: {e2}")
            results.append((fname, "error", err_msg))

    # 요약 CSV 저장
    summary_path = os.path.join(
        BASE_DIR,
        f"vol_band_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
    )
    rows = []
    for fname, flag, info in results:
        row = {"file": fname, "flag": flag}
        if isinstance(info, dict):
            row.update(
                {
                    "last": info.get("last"),
                    "mu": info.get("mu"),
                    "sd": info.get("sd"),
                    "upper": info.get("upper"),
                    "lower": info.get("lower"),
                    "zscore": info.get("zscore"),
                    "thr_yellow": info.get("thr_yellow"),
                    "thr_orange": info.get("thr_orange"),
                    "thr_red": info.get("thr_red"),
                    "level": info.get("level"),
                    "reason": info.get("reason"),
                }
            )
        elif isinstance(info, str):
            row["error"] = info
        rows.append(row)
    pd.DataFrame(rows).to_csv(summary_path, index=False, encoding="utf-8-sig")
    print(f"\n요약 저장: {summary_path}")


if __name__ == "__main__":
    main()


# -*- coding: utf-8 -*-
"""
fnnews 기사 1건 샘플 처리 코드
- URL: https://www.fnnews.com/news/202511051823464896
- Selenium으로 body.innerText 추출 → 텍스트 정제 → OpenAI GPT 요약
"""

import time
import re
from pathlib import Path
from typing import List, Optional

# Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException, TimeoutException, NoSuchElementException

# OpenAI 신/구 SDK 자동 감지용
# pip install openai 필요


# =========================
# 환경 설정
# =========================

# 샘플 기사 URL
ARTICLE_URL = "https://www.fnnews.com/news/202511051823464896"

# OpenAI 키 파일 경로 (팀장님 기존 경로 그대로 사용)
KEY_PATH = Path(r"C:\Users\amongpapa\lm\keys\open.txt")

# 크롬 실행 경로 / 프로필 (필요시 수정해서 사용)
CHROME_PATH = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")
CHROME_PROFILE = "Profile 1"  # 예: 'Profile 1', 'Default' 등

# 페이지 로딩 대기시간(초) — 느리면 8~10초 정도로 늘려주세요
PAGE_LOAD_WAIT = 8.0

# OpenAI 모델명
MODEL_ID = "gpt-4o-mini"  # 필요시 다른 모델로 변경


# =========================
# OpenAI 신/구 SDK 자동 호환
# =========================

def load_api_key(path: Path) -> str:
    """
    텍스트 파일에서 OpenAI API 키를 읽어오는 함수.
    - 팀장님 환경처럼 open.txt에 키 한 줄만 들어있다고 가정.
    """
    key = path.read_text(encoding="utf-8").strip()
    if not key:
        raise ValueError("API 키가 비어 있습니다.")
    return key


class ChatClient:
    """
    신/구 OpenAI SDK 자동 감지용 래퍼.
    - .chat(messages, model=..., temperature=...) → str(답변 텍스트)
    """
    def __init__(self, api_key: str):
        self.mode = None
        try:
            # 신버전(openai>=1.x)
            from openai import OpenAI  # type: ignore
            self._client = OpenAI(api_key=api_key)
            self.mode = "new"
        except Exception:
            # 구버전(openai<1.x)
            import openai  # type: ignore
            openai.api_key = api_key
            self._client = openai
            self.mode = "old"

    def chat(self, messages: List[dict], model: str, temperature: float = 0.2) -> str:
        """
        messages: [{"role": "system"/"user"/"assistant", "content": "..."}]
        model   : OpenAI 모델명 (예: "gpt-4o-mini")
        반환값  : assistant 메시지 content 문자열
        """
        if self.mode == "new":
            res = self._client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res.choices[0].message.content.strip()
        else:
            res = self._client.ChatCompletion.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res["choices"][0]["message"]["content"].strip()


# =========================
# 텍스트 정리 & 프롬프트 구성
# =========================

def clean_text(text: str) -> str:
    """
    기사 본문 텍스트 간단 정제:
    - 캐리지리턴 제거 → 줄바꿈 통일
    - 3줄 이상 연속 개행 → 2줄로 줄이기
    - 공백/탭 여러 개 → 1칸
    - '쿠키', '이용약관', '저작권', '광고문의' 등 전형적인 잡음 문장 제거
    """
    # 줄바꿈 정리
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    # 공백 정리
    text = re.sub(r"[ \t]+", " ", text)

    # 제거 패턴
    drop_patterns = [
        r"쿠키(를|에) 사용",
        r"이용약관",
        r"개인정보",
        r"구독",
        r"광고문의",
        r"무단전재",
        r"저작권",
    ]

    cleaned_lines = []
    for ln in text.splitlines():
        if any(re.search(pat, ln, re.IGNORECASE) for pat in drop_patterns):
            continue
        cleaned_lines.append(ln.strip())

    return "\n".join(cleaned_lines).strip()


PROMPT_TEMPLATE = (
    "기사 요약 – 서술형 보고서\n"
    "[역할] 당신은 은행 리스크관리 보고서 에디터다.\n"
    "[목표] 아래 기사 원문을 바탕으로 한국어 서술형 요약을 작성한다. 총 길이 {2000}자 이내.\n"
    "[입력] {기사 원문}\n"
    "[출력 지침]\n"
    "- 제목: 한 줄\n"
    "- 요약: 3문장\n"
    "- 본문: 사실→원인→영향→전망 순으로 3–6단락.\n"
    "- 숫자·기관명·지명은 가능한 한 정확히 유지.\n"
    "- 마지막 문장에 (신뢰도 0.0–1.0; 근거 요약) 형식으로 작성.\n"
    "[금지] 광고문구, 과장된 표현, 필요 없는 안내문, 표/불릿/JSON 출력.\n"
    "[실행] 위 지침을 적용해 결과만 출력한다.\n"
)

def build_prompt(article_text: str, max_chars: int = 2000) -> str:
    """
    GPT에게 넘길 프롬프트 생성.
    - 기사 텍스트가 너무 길면 8000자까지만 사용 (과도한 토큰 방지)
    - {2000}, {기사 원문} 자리에 실제 값 채워 넣기
    """
    article_text = article_text[:8000]
    return (
        PROMPT_TEMPLATE
        .replace("{2000}", str(max_chars))
        .replace("{기사 원문}", article_text)
    )


def summarize_article(client: ChatClient, article_text: str, model: str = MODEL_ID) -> str:
    """
    정제된 기사 텍스트 → GPT 요약 호출 함수.
    - system 메시지로 '은행 리스크관리 보조원' 역할 부여
    - user 메시지에 빌드된 프롬프트 입력
    """
    sys_msg = (
        "너는 은행 리스크관리 보조원이며, 출력은 반드시 한국어로 작성한다. "
        "광고/구독/저작권/추천기사/댓글 등 불필요 텍스트는 제거하고, "
        "요구된 형식만 간결하게 작성한다."
    )
    messages = [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": build_prompt(article_text, max_chars=2000)},
    ]
    return client.chat(messages, model=model, temperature=0.2)


# =========================
# Selenium 드라이버 구성/해제
# =========================

def build_driver(download_dir: Optional[Path] = None) -> webdriver.Chrome:
    """
    Selenium Chrome 드라이버 생성 함수.
    - CHROME_PATH, CHROME_PROFILE 사용
    - 필요시 user-data-dir, 프록시 등 옵션 추가 가능
    """
    chrome_options = webdriver.ChromeOptions()

    # 크롬 실행파일 지정 (설치 위치에 따라 수정)
    if CHROME_PATH.exists():
        chrome_options.binary_location = str(CHROME_PATH)

    # 팀장님이 평소 쓰는 크롬 프로필을 그대로 쓰고 싶을 때 (선택)
    chrome_options.add_argument(f"--profile-directory={CHROME_PROFILE}")
    # 예: 회사 환경에서 프록시/보안 설정이 프로필에 묶여 있는 경우 유리

    # 기타 옵션들
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")

    if download_dir:
        prefs = {
            "download.default_directory": str(download_dir),
            "download.prompt_for_download": False,
            "directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        chrome_options.add_experimental_option("prefs", prefs)

    # chromedriver가 PATH에 있다고 가정
    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(30)
    return driver


def safe_quit(driver: webdriver.Chrome):
    """드라이버 종료 시 예외가 나더라도 무시하고 안전하게 종료."""
    try:
        driver.quit()
    except Exception:
        pass


# =========================
# 페이지 텍스트 추출 함수
# =========================

def get_visible_text(driver: webdriver.Chrome, url: str) -> str:
    """
    주어진 URL을 열고, document.body.innerText를 가져오는 함수.
    - PAGE_LOAD_WAIT 동안 대기 후 body 텍스트 추출
    - 예외 발생 시 빈 문자열 반환
    - 디버깅용으로 len(text) 출력
    """
    try:
        print(f"[INFO] 페이지 열기: {url}")
        driver.get(url)
        time.sleep(PAGE_LOAD_WAIT)

        try:
            text = driver.execute_script(
                "return document.body ? document.body.innerText : '';"
            )
        except Exception:
            try:
                text = driver.find_element(By.TAG_NAME, "body").text
            except NoSuchElementException:
                text = ""

        text = text or ""
        print(f"[DEBUG] 추출된 텍스트 길이: {len(text)}")
        return text

    except (WebDriverException, TimeoutException) as e:
        print(f"[ERR] 페이지 로딩 실패: {e}")
        return ""


# =========================
# 메인 실행부
# =========================

def main():
    # 1) OpenAI 클라이언트 준비
    api_key = load_api_key(KEY_PATH)
    client = ChatClient(api_key=api_key)

    # 2) Selenium 드라이버 띄우기
    driver = build_driver(download_dir=None)

    try:
        # 3) 기사 본문(raw 텍스트) 가져오기
        raw_text = get_visible_text(driver, ARTICLE_URL)
        if not raw_text:
            print("[WARN] 본문 텍스트를 가져오지 못했습니다. (회사망/프록시 문제일 수 있음)")
            return

        # 4) 텍스트 정제
        cleaned = clean_text(raw_text)
        print(f"[DEBUG] 정제 후 텍스트 길이: {len(cleaned)}")

        # 5) GPT 요약 생성
        summary = summarize_article(client, cleaned, model=MODEL_ID)

        # 6) 결과 출력
        print("\n==================== [기사 요약 결과] ====================")
        print(summary)
        print("=========================================================\n")

    finally:
        # 7) 드라이버 종료
        safe_quit(driver)


if __name__ == "__main__":
    main()


# -*- coding: utf-8 -*-
"""
keyword_news_bf.xlsx 전체 테스트용 (디버깅 버전)
- 각 행의 URL → 기사 본문 추출 → GPT 요약 → news 컬럼 채우기
- 각 행마다 디버깅 로그 출력
"""

import time
import re
from pathlib import Path
from typing import List, Optional

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException, TimeoutException, NoSuchElementException

# ===== 환경 설정 =====
DIR_PATH   = Path(r"C:\Users\amongpapa\chartup\go_scen\data\news\bigkinds\daily_news")
SOURCE_BF  = DIR_PATH / "keyword_news_bf.xlsx"
KEY_PATH   = Path(r"C:\Users\amongpapa\lm\keys\open.txt")

CHROME_PATH    = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")
CHROME_PROFILE = "Profile 1"
PAGE_LOAD_WAIT = 8.0
MODEL_ID       = "gpt-4o-mini"


# ===== OpenAI 신/구 SDK 래퍼 =====
def load_api_key(path: Path) -> str:
    key = path.read_text(encoding="utf-8").strip()
    if not key:
        raise ValueError("API 키가 비어 있습니다.")
    return key


class ChatClient:
    def __init__(self, api_key: str):
        self.mode = None
        try:
            from openai import OpenAI  # 신버전
            self._client = OpenAI(api_key=api_key)
            self.mode = "new"
        except Exception:
            import openai              # 구버전
            openai.api_key = api_key
            self._client = openai
            self.mode = "old"

    def chat(self, messages: List[dict], model: str, temperature: float = 0.2) -> str:
        if self.mode == "new":
            res = self._client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res.choices[0].message.content.strip()
        else:
            res = self._client.ChatCompletion.create(
                model=model,
                messages=messages,
                temperature=temperature,
            )
            return res["choices"][0]["message"]["content"].strip()


# ===== 텍스트 정리 / 프롬프트 =====
def clean_text(text: str) -> str:
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)

    drop_patterns = [
        r"쿠키(를|에) 사용",
        r"이용약관",
        r"개인정보",
        r"구독",
        r"광고문의",
        r"무단전재",
        r"저작권",
    ]
    cleaned_lines = []
    for ln in text.splitlines():
        if any(re.search(pat, ln, re.IGNORECASE) for pat in drop_patterns):
            continue
        cleaned_lines.append(ln.strip())
    return "\n".join(cleaned_lines).strip()


PROMPT_TEMPLATE = (
    "기사 요약 – 서술형 보고서\n"
    "[역할] 당신은 은행 리스크관리 보고서 에디터다.\n"
    "[목표] 아래 기사 원문을 바탕으로 한국어 서술형 요약을 작성한다. 총 길이 {2000}자 이내.\n"
    "[입력] {기사 원문}\n"
    "[출력 지침]\n"
    "- 제목: 한 줄\n"
    "- 요약: 3문장\n"
    "- 본문: 사실→원인→영향→전망 순으로 3–6단락.\n"
    "- 마지막 문장에 (신뢰도 0.0–1.0; 근거 요약) 형식으로 작성.\n"
    "[금지] 광고문구, 과장된 표현, 필요 없는 안내문, 표/불릿/JSON 출력.\n"
    "[실행] 위 지침을 적용해 결과만 출력한다.\n"
)

def build_prompt(article_text: str, max_chars: int = 2000) -> str:
    article_text = article_text[:8000]
    return (
        PROMPT_TEMPLATE
        .replace("{2000}", str(max_chars))
        .replace("{기사 원문}", article_text)
    )

def summarize_article(client: ChatClient, article_text: str, model: str = MODEL_ID) -> str:
    sys_msg = (
        "너는 은행 리스크관리 보조원이며, 출력은 반드시 한국어로 작성한다. "
        "광고/구독/저작권/추천기사/댓글 등 불필요 텍스트는 제거하고, "
        "요구된 형식만 간결하게 작성한다."
    )
    messages = [
        {"role": "system", "content": sys_msg},
        {"role": "user", "content": build_prompt(article_text, max_chars=2000)},
    ]
    return client.chat(messages, model=model, temperature=0.2)


# ===== Selenium 드라이버 =====
def build_driver(download_dir: Optional[Path] = None) -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    if CHROME_PATH.exists():
        chrome_options.binary_location = str(CHROME_PATH)

    chrome_options.add_argument(f"--profile-directory={CHROME_PROFILE}")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")

    if download_dir:
        prefs = {
            "download.default_directory": str(download_dir),
            "download.prompt_for_download": False,
            "directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(30)
    return driver

def safe_quit(driver: webdriver.Chrome):
    try:
        driver.quit()
    except Exception:
        pass


# ===== news 채움 여부 판정 =====
def should_fill(val) -> bool:
    """
    news 칼럼이 비어 있으면 True, 뭔가 들어있으면 False.
    (NaN / None / 빈 문자열 → True)
    """
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    if isinstance(val, str) and not val.strip():
        return True
    return False


# ===== 메인 =====
def main():
    api_key = load_api_key(KEY_PATH)
    client  = ChatClient(api_key=api_key)

    if not SOURCE_BF.exists():
        print(f"[ERR] 소스 파일 없음: {SOURCE_BF}")
        return

    df = pd.read_excel(SOURCE_BF, engine="openpyxl")
    print(f"[INFO] 로드: {SOURCE_BF.name}, shape={df.shape}")

    if "URL" not in df.columns:
        print("[ERR] 'URL' 컬럼이 없습니다.")
        return
    if "news" not in df.columns:
        df["news"] = pd.NA

    driver = build_driver(download_dir=None)

    try:
        updated = 0
        for idx, row in df.iterrows():
            url = str(row.get("URL", "")).strip()
            news_val = row.get("news", None)
            fill_flag = should_fill(news_val)

            print(f"\n[DEBUG] row={idx}, fill?={fill_flag}, news={repr(news_val)}")
            print(f"[DEBUG] URL={url}")

            if not url.startswith("http"):
                print("[SKIP] http/https 아님")
                continue
            if not fill_flag:
                print("[SKIP] news가 이미 채워져 있음")
                continue

            raw = get_visible_text(driver, url)
            if not raw or len(raw) < 30:
                print(f"[WARN] 텍스트가 너무 짧음(len={len(raw)})")
                continue

            cleaned = clean_text(raw)
            print(f"[DEBUG] 정제 후 텍스트 길이: {len(cleaned)}")

            try:
                summary = summarize_article(client, cleaned, model=MODEL_ID)
            except Exception as e:
                print(f"[ERR] GPT 요약 실패: {e}")
                continue

            df.at[idx, "news"] = summary
            updated += 1
            print(f"[OK] row={idx} 요약 완료")

            time.sleep(0.5)  # 과금/속도 조절용

        df.to_excel(SOURCE_BF, index=False, engine="openpyxl")
        print(f"\n[OK] 저장 완료: {SOURCE_BF.name}, updated={updated}")

    finally:
        safe_quit(driver)


def get_visible_text(driver: webdriver.Chrome, url: str) -> str:
    """
    단일 URL에서 body.innerText를 추출하고 길이를 출력.
    """
    try:
        print(f"[INFO] 페이지 열기: {url}")
        driver.get(url)
        time.sleep(PAGE_LOAD_WAIT)

        try:
            text = driver.execute_script(
                "return document.body ? document.body.innerText : '';"
            )
        except Exception:
            try:
                text = driver.find_element(By.TAG_NAME, "body").text
            except NoSuchElementException:
                text = ""

        text = text or ""
        print(f"[DEBUG] 추출된 텍스트 길이: {len(text)}")
        return text

    except (WebDriverException, TimeoutException) as e:
        print(f"[ERR] 페이지 로딩 실패: {e}")
        return ""


if __name__ == "__main__":
    main()


